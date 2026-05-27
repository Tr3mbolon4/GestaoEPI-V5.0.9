from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Request
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from database import connect_db, close_db, get_db
from schemas import *
from auth import *
from datetime import datetime, timezone, timedelta
from typing import List, Optional
from pathlib import Path
from bson import ObjectId
import os
import shutil
import logging
import base64
import json
import io
import re
import unicodedata

# Para importação de Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Para geração de PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from services.face_recognition_service import FaceRecognitionService

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent
UPLOAD_DIR = ROOT_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)

# Dias para expiração de senha
PASSWORD_EXPIRY_DAYS = 30
face_recognition_service = FaceRecognitionService()

app = FastAPI(title='Cipolatti API')
api_router = APIRouter(prefix='/api')

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=['*'],
    allow_methods=['*'],
    allow_headers=['*'],
    expose_headers=['*'],
)

# Middleware para adicionar headers de cache-control e CORS em arquivos estáticos
@app.middleware("http")
async def add_cache_headers(request, call_next):
    response = await call_next(request)
    
    # Para arquivos de upload, adicionar headers de cache e CORS
    if '/uploads/' in request.url.path or '/api/uploads/' in request.url.path:
        response.headers['Cache-Control'] = 'public, max-age=86400'  # 1 dia
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        response.headers['X-Content-Type-Options'] = 'nosniff'
    
    return response

# Servir arquivos de upload - DEVE ter prefixo /api para funcionar com ingress
app.mount('/api/uploads', StaticFiles(directory=str(UPLOAD_DIR)), name='api_uploads')
app.mount('/uploads', StaticFiles(directory=str(UPLOAD_DIR)), name='uploads')

@app.on_event('startup')
async def startup_event():
    await connect_db()
    from seed import seed_database
    await seed_database()
    db = await get_db()
    await face_recognition_service.initialize_cache(db)

@app.on_event('shutdown')
async def shutdown_event():
    await close_db()

def doc_to_response(doc, id_field='id'):
    if doc is None:
        return None
    result = {k: v for k, v in doc.items() if k != '_id'}
    result[id_field] = str(doc['_id'])
    return result

def parse_datetime_safe(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace('Z', '+00:00'))
        except Exception:
            return None
    return None

def parse_int_safe(value, default=0):
    if value is None or value == '':
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default

def get_variation_stock_quantity(raw_variation):
    quantity = raw_variation.get('quantidade_estoque')
    if quantity is None:
        quantity = raw_variation.get('current_stock')
    return parse_int_safe(quantity, 0)

def ensure_aware_datetime(value):
    parsed = parse_datetime_safe(value)
    if not parsed:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed

def format_datetime_safe(value, fmt='%d/%m/%Y %H:%M', fallback='-'):
    parsed = ensure_aware_datetime(value)
    return parsed.strftime(fmt) if parsed else fallback

def format_facial_template_response(doc):
    if doc is None:
        return None
    return {
        "id": str(doc["_id"]),
        "employee_id": str(doc.get("employee_id")) if doc.get("employee_id") else None,
        "descriptor": doc.get("descriptor"),
        "pose_label": doc.get("pose_label"),
        "quality_score": doc.get("quality_score"),
        "detection_score": doc.get("detection_score"),
        "created_by": doc.get("created_by"),
        "created_by_name": doc.get("created_by_name"),
        "created_at": doc.get("created_at")
    }

def biometric_quality_label(quality_score: Optional[float]) -> str:
    score = float(quality_score or 0)
    if score >= 0.85:
        return "Excelente"
    if score >= 0.70:
        return "Boa"
    if score >= 0.50:
        return "Regular"
    return "Ruim"

def has_registered_facial_template(template: dict) -> bool:
    descriptor = template.get("descriptor")
    if descriptor is None:
        return False
    if isinstance(descriptor, str):
        stripped = descriptor.strip()
        if not stripped:
            return False
        try:
            parsed = json.loads(stripped)
            if isinstance(parsed, list):
                return len(parsed) > 0
        except Exception:
            pass
        return True
    if isinstance(descriptor, (list, tuple)):
        return len(descriptor) > 0
    return True

def build_biometric_summary(employee: dict, templates: List[dict]) -> dict:
    registered_templates = [
        template for template in templates
        if has_registered_facial_template(template)
    ]
    compatible_templates = [
        template for template in templates
        if face_recognition_service.is_descriptor_compatible(template.get("descriptor"))
    ]
    count = len(registered_templates)
    if count >= 3:
        status_key = "registered"
        status_label = "Cadastrada"
    elif count > 0:
        status_key = "incomplete"
        status_label = "Incompleta"
    else:
        status_key = "missing"
        status_label = "Nao cadastrada"

    dates = [
        ensure_aware_datetime(template.get("created_at"))
        for template in registered_templates
        if ensure_aware_datetime(template.get("created_at"))
    ]
    quality_scores = [float(template.get("quality_score") or 0) for template in registered_templates]
    avg_quality = round(sum(quality_scores) / len(quality_scores), 4) if quality_scores else 0
    latest_template = sorted(
        registered_templates,
        key=lambda template: ensure_aware_datetime(template.get("created_at")) or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True
    )[0] if registered_templates else None

    return {
        "status": status_key,
        "status_label": status_label,
        "templates_count": count,
        "total_templates_count": len(templates),
        "compatible_templates_count": len(compatible_templates),
        "legacy_templates_count": max(count - len(compatible_templates), 0),
        "first_enrolled_at": min(dates) if dates else None,
        "last_updated_at": max(dates) if dates else None,
        "created_by": latest_template.get("created_by") if latest_template else None,
        "created_by_name": latest_template.get("created_by_name") if latest_template else None,
        "model": "InsightFace",
        "has_consent": employee.get("facial_consent", False),
        "consent_date": employee.get("facial_consent_date"),
        "quality_score": avg_quality,
        "quality_label": biometric_quality_label(avg_quality),
    }

async def get_biometric_summaries(db, employees: List[dict]) -> dict:
    employee_ids = [str(employee["_id"]) for employee in employees]
    object_ids = [employee["_id"] for employee in employees]
    templates = await db.facial_templates.find({
        "employee_id": {"$in": employee_ids + object_ids}
    }).to_list(10000)
    templates_by_employee = {}
    for template in templates:
        templates_by_employee.setdefault(str(template.get("employee_id")), []).append(template)
    return {
        str(employee["_id"]): build_biometric_summary(
            employee,
            templates_by_employee.get(str(employee["_id"]), [])
        )
        for employee in employees
    }

def check_password_expired(user):
    """Verifica se a senha expirou (mais de 30 dias)"""
    password_changed_at = ensure_aware_datetime(user.get('password_changed_at'))
    if not password_changed_at:
        return False

    expiry_date = password_changed_at + timedelta(days=PASSWORD_EXPIRY_DAYS)
    return datetime.now(timezone.utc) > expiry_date

# Permissões por perfil
ROLE_PERMISSIONS = {
    'admin': ['all'],
    'gestor': ['dashboard', 'entrega', 'colaboradores', 'empresas', 'epis', 'fornecedores', 'kits'],
    'rh': ['dashboard', 'colaboradores', 'colaboradores_full', 'empresas', 'usuarios'],
    'seguranca_trabalho': ['dashboard', 'epis', 'fornecedores', 'kits', 'colaboradores_list'],
    'almoxarifado': ['dashboard', 'entrega', 'colaboradores_list']
}

def can_view_sensitive_data(role):
    """Verifica se o perfil pode ver dados sensíveis (CPF, RG, etc)"""
    return role in ['admin', 'gestor', 'rh']

def can_manage_users(role):
    """Verifica se pode gerenciar usuários"""
    return role in ['admin', 'rh']

def can_deliver_epi(role):
    """Verifica se pode fazer entregas de EPI"""
    return role in ['admin', 'gestor', 'almoxarifado']

def can_manage_epis(role):
    """Verifica se pode gerenciar EPIs - inclui almoxarifado"""
    return role in ['admin', 'gestor', 'seguranca_trabalho', 'almoxarifado']

def can_manage_employees(role):
    """Verifica se pode cadastrar/editar colaboradores"""
    return role in ['admin', 'gestor', 'rh']

# ===================== AUTH =====================

@api_router.get('/')
async def root():
    return {'message': 'Cipolatti API'}

@api_router.get('/health')
async def health_check():
    """Endpoint de verificação de saúde da API"""
    # Verificar conexão com MongoDB
    db_status = 'unknown'
    try:
        db = await get_db()
        await db.command('ping')
        db_status = 'connected'
    except Exception as e:
        db_status = f'error: {str(e)}'
    
    # Verificar pasta de uploads
    uploads_status = 'ok' if os.path.isdir(UPLOAD_DIR) else 'missing'
    
    # Contar arquivos de upload
    upload_count = 0
    if os.path.isdir(UPLOAD_DIR):
        for root_dir, dirs, files in os.walk(UPLOAD_DIR):
            upload_count += len(files)
    
    return {
        'status': 'healthy' if db_status == 'connected' else 'degraded',
        'database': db_status,
        'uploads_dir': uploads_status,
        'upload_files_count': upload_count,
        'backend_url': os.environ.get('BACKEND_URL', 'not_set'),
        'timestamp': datetime.now(timezone.utc).isoformat()
    }

@api_router.post('/auth/login', response_model=TokenResponse)
async def login(request: LoginRequest):
    db = await get_db()
    
    # Verificar se o identificador é email ou username
    identifier = request.username.strip()
    if '@' in identifier:
        # Buscar por email
        user = await db.users.find_one({"email": identifier.lower()})
    else:
        # Buscar por username
        user = await db.users.find_one({"username": identifier})
    
    if not user or not verify_password(request.password, user['hashed_password']):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail='Credenciais inválidas')
    
    if not user.get('is_active', True):
        raise HTTPException(status_code=400, detail='Usuário inativo')
    
    license_doc = await db.panel_license.find_one({})
    if license_doc:
        now = datetime.now(timezone.utc)
        expires_at = ensure_aware_datetime(license_doc.get('expires_at'))
        if not expires_at:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail='Licenca invalida')
        if now > expires_at and user['role'] != 'admin':
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail='Licença expirada')
    
    password_expired = check_password_expired(user)
    
    access_token = create_access_token(data={'sub': user['username'], 'role': user['role']})
    
    return TokenResponse(
        access_token=access_token,
        token_type='bearer',
        must_change_password=user.get('must_change_password', False),
        password_expired=password_expired,
        role=UserRole(user['role']),
        is_primary_admin=user.get('is_primary_admin', False)
    )

@api_router.post('/auth/change-password')
async def change_password(request: ChangePasswordRequest, current_user: dict = Depends(get_current_user)):
    db = await get_db()
    if not verify_password(request.old_password, current_user['hashed_password']):
        raise HTTPException(status_code=400, detail='Senha antiga incorreta')
    
    await db.users.update_one(
        {"_id": ObjectId(current_user['id'])},
        {"$set": {
            "hashed_password": get_password_hash(request.new_password), 
            "must_change_password": False,
            "password_changed_at": datetime.now(timezone.utc)
        }}
    )
    return {'message': 'Senha alterada com sucesso'}

@api_router.get('/auth/me', response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        id=current_user['id'],
        username=current_user['username'],
        email=current_user['email'],
        role=UserRole(current_user['role']),
        is_active=current_user.get('is_active', True),
        must_change_password=current_user.get('must_change_password', False),
        password_changed_at=current_user.get('password_changed_at'),
        employee_id=current_user.get('employee_id'),
        created_at=current_user.get('created_at', datetime.now(timezone.utc))
    )

# ===================== USERS =====================

@api_router.get('/users', response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(require_role('admin', 'rh'))):
    db = await get_db()
    users = await db.users.find({}).to_list(1000)
    return [UserResponse(**doc_to_response(u)) for u in users]

@api_router.post('/users', response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def create_user(user_data: UserCreate, current_user: dict = Depends(require_role('admin', 'rh'))):
    db = await get_db()
    
    # RH não pode criar admins
    if current_user['role'] == 'rh' and user_data.role == UserRole.ADMIN:
        raise HTTPException(status_code=403, detail='Sem permissão para criar administradores')
    
    # Normalizar email para lowercase
    email_normalized = user_data.email.lower().strip() if user_data.email else None
    
    existing = await db.users.find_one({"$or": [
        {"username": user_data.username}, 
        {"email": email_normalized}
    ]})
    if existing:
        if existing.get('username') == user_data.username:
            raise HTTPException(status_code=400, detail='Nome de usuário já existe')
        else:
            raise HTTPException(status_code=400, detail='E-mail já está em uso')
    
    new_user = {
        "username": user_data.username,
        "email": email_normalized,
        "hashed_password": get_password_hash(user_data.password),
        "role": user_data.role.value,
        "employee_id": user_data.employee_id,
        "must_change_password": True,
        "is_active": True,
        "password_changed_at": datetime.now(timezone.utc),
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    result = await db.users.insert_one(new_user)
    new_user['_id'] = result.inserted_id
    return UserResponse(**doc_to_response(new_user))

@api_router.patch('/users/{user_id}', response_model=UserResponse)
async def update_user(user_id: str, user_data: UserUpdate, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    update_data = {k: v for k, v in user_data.model_dump(exclude_unset=True).items()}
    if 'role' in update_data:
        update_data['role'] = update_data['role'].value
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    result = await db.users.find_one_and_update(
        {"_id": ObjectId(user_id)}, {"$set": update_data}, return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail='Usuário não encontrado')
    return UserResponse(**doc_to_response(result))

@api_router.delete('/users/{user_id}')
async def delete_user(user_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    result = await db.users.delete_one({"_id": ObjectId(user_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Usuário não encontrado')
    return {'message': 'Usuário excluído'}

class ResetPasswordRequest(BaseModel):
    new_password: str

@api_router.post('/users/{user_id}/reset-password')
async def reset_user_password(user_id: str, request: ResetPasswordRequest, current_user: dict = Depends(require_role('admin'))):
    """Permite ao administrador redefinir a senha de qualquer usuário"""
    db = await get_db()
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail='Usuário não encontrado')
    
    # Validar complexidade da senha
    password = request.new_password
    import re
    if len(password) < 8:
        raise HTTPException(status_code=400, detail='A senha deve ter no mínimo 8 caracteres')
    if not re.search(r'[A-Z]', password):
        raise HTTPException(status_code=400, detail='A senha deve conter pelo menos uma letra maiúscula')
    if not re.search(r'[a-z]', password):
        raise HTTPException(status_code=400, detail='A senha deve conter pelo menos uma letra minúscula')
    if not re.search(r'\d', password):
        raise HTTPException(status_code=400, detail='A senha deve conter pelo menos um número')
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
        raise HTTPException(status_code=400, detail='A senha deve conter pelo menos um caractere especial')
    
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {
            "hashed_password": get_password_hash(password),
            "must_change_password": True,
            "password_changed_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {'message': 'Senha redefinida com sucesso'}

# ===================== COMPANIES =====================

@api_router.get('/companies', response_model=List[CompanyResponse])
async def get_companies(current_user: dict = Depends(require_role('admin', 'gestor', 'rh'))):
    db = await get_db()
    companies = await db.companies.find({}).to_list(1000)
    return [CompanyResponse(**doc_to_response(c)) for c in companies]

@api_router.post('/companies', response_model=CompanyResponse, status_code=status.HTTP_201_CREATED)
async def create_company(company_data: CompanyCreate, current_user: dict = Depends(require_role('admin', 'gestor', 'rh'))):
    logger.info(f"📝 Recebendo dados de empresa: {company_data.model_dump()}")
    db = await get_db()
    existing = await db.companies.find_one({"cnpj": company_data.cnpj})
    if existing:
        logger.error(f"❌ CNPJ já cadastrado: {company_data.cnpj}")
        raise HTTPException(status_code=400, detail='CNPJ já cadastrado')
    
    new_company = {**company_data.model_dump(), "created_at": datetime.now(timezone.utc), "updated_at": datetime.now(timezone.utc)}
    logger.info(f"✅ Inserindo empresa: {new_company}")
    result = await db.companies.insert_one(new_company)
    new_company['_id'] = result.inserted_id
    logger.info(f"✅ Empresa criada com ID: {result.inserted_id}")
    return CompanyResponse(**doc_to_response(new_company))

@api_router.get('/companies/{company_id}', response_model=CompanyResponse)
async def get_company(company_id: str, current_user: dict = Depends(require_role('admin', 'gestor', 'rh'))):
    db = await get_db()
    company = await db.companies.find_one({"_id": ObjectId(company_id)})
    if not company:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    return CompanyResponse(**doc_to_response(company))

@api_router.patch('/companies/{company_id}', response_model=CompanyResponse)
async def update_company(company_id: str, company_data: CompanyUpdate, current_user: dict = Depends(require_role('admin', 'gestor', 'rh'))):
    db = await get_db()
    update_data = {k: v for k, v in company_data.model_dump(exclude_unset=True).items()}
    update_data['updated_at'] = datetime.now(timezone.utc)
    result = await db.companies.find_one_and_update(
        {"_id": ObjectId(company_id)}, {"$set": update_data}, return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    return CompanyResponse(**doc_to_response(result))

@api_router.delete('/companies/{company_id}')
async def delete_company(company_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    result = await db.companies.delete_one({"_id": ObjectId(company_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    return {'message': 'Empresa excluída'}

# ===================== EMPLOYEES =====================

@api_router.get('/employees')
async def get_employees(current_user: dict = Depends(get_current_user), search: Optional[str] = None, company_id: Optional[str] = None):
    db = await get_db()
    query = {}
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"cpf": {"$regex": search, "$options": "i"}},
            {"registration_number": {"$regex": search, "$options": "i"}}
        ]
    if company_id:
        query["company_id"] = company_id
    
    employees = await db.employees.find(query).to_list(1000)
    biometric_summaries = await get_biometric_summaries(db, employees)
    
    # Perfis que não podem ver dados sensíveis
    if not can_view_sensitive_data(current_user['role']):
        result = []
        for employee in employees:
            item = EmployeePublicResponse(**doc_to_response(employee)).model_dump()
            summary = biometric_summaries.get(str(employee['_id']))
            item["biometric"] = summary
            item["biometric_status"] = summary["status"] if summary else "missing"
            item["biometric_templates_count"] = summary["templates_count"] if summary else 0
            result.append(item)
        return result
    
    result = []
    for employee in employees:
        item = EmployeeResponse(**doc_to_response(employee)).model_dump()
        summary = biometric_summaries.get(str(employee['_id']))
        item["biometric"] = summary
        item["biometric_status"] = summary["status"] if summary else "missing"
        item["biometric_templates_count"] = summary["templates_count"] if summary else 0
        result.append(item)
    return result

@api_router.post('/employees', response_model=EmployeeResponse, status_code=status.HTTP_201_CREATED)
async def create_employee(employee_data: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão para cadastrar colaboradores')
    
    db = await get_db()
    existing = await db.employees.find_one({"cpf": employee_data.cpf})
    if existing:
        raise HTTPException(status_code=400, detail='CPF já cadastrado')
    
    new_employee = {**employee_data.model_dump(), "created_at": datetime.now(timezone.utc), "updated_at": datetime.now(timezone.utc)}
    result = await db.employees.insert_one(new_employee)
    new_employee['_id'] = result.inserted_id
    return EmployeeResponse(**doc_to_response(new_employee))

@api_router.get('/employees/{employee_id}')
async def get_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    db = await get_db()
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    templates = await db.facial_templates.find({
        "employee_id": {"$in": [employee_id, ObjectId(employee_id)]}
    }).to_list(100)
    biometric_summary = build_biometric_summary(employee, templates)

    if not can_view_sensitive_data(current_user['role']):
        item = EmployeePublicResponse(**doc_to_response(employee)).model_dump()
        item["biometric"] = biometric_summary
        item["biometric_status"] = biometric_summary["status"]
        item["biometric_templates_count"] = biometric_summary["templates_count"]
        return item

    item = EmployeeResponse(**doc_to_response(employee)).model_dump()
    item["biometric"] = biometric_summary
    item["biometric_status"] = biometric_summary["status"]
    item["biometric_templates_count"] = biometric_summary["templates_count"]
    return item

@api_router.patch('/employees/{employee_id}', response_model=EmployeeResponse)
async def update_employee(employee_id: str, employee_data: EmployeeUpdate, current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão para editar colaboradores')
    
    db = await get_db()
    update_data = {k: v for k, v in employee_data.model_dump(exclude_unset=True).items()}
    update_data['updated_at'] = datetime.now(timezone.utc)
    result = await db.employees.find_one_and_update(
        {"_id": ObjectId(employee_id)}, {"$set": update_data}, return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    return EmployeeResponse(**doc_to_response(result))

@api_router.delete('/employees/{employee_id}')
async def delete_employee(employee_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    result = await db.employees.delete_one({"_id": ObjectId(employee_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    return {'message': 'Colaborador excluído'}

@api_router.post('/employees/{employee_id}/photo')
async def upload_employee_photo(employee_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão')
    
    db = await get_db()
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    
    file_ext = Path(file.filename).suffix
    file_name = f'employee_{employee_id}_{datetime.now(timezone.utc).timestamp()}{file_ext}'
    file_path = UPLOAD_DIR / 'employees' / file_name
    file_path.parent.mkdir(exist_ok=True, parents=True)
    
    with file_path.open('wb') as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    photo_path = f'/uploads/employees/{file_name}'
    await db.employees.update_one({"_id": ObjectId(employee_id)}, {"$set": {"photo_path": photo_path}})
    return {'photo_path': photo_path}

# ===================== IMPORTAÇÃO/EXPORTAÇÃO =====================

@api_router.get('/employees/export/excel')
async def export_employees_excel(current_user: dict = Depends(get_current_user)):
    """Exporta colaboradores para Excel"""
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Permissão insuficiente')
    
    db = await get_db()
    employees = await db.employees.find().to_list(5000)
    companies = {str(c['_id']): c['legal_name'] async for c in db.companies.find()}
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores"
    
    # Cabeçalho
    headers = ['Nome Completo', 'CPF', 'RG', 'Matrícula', 'Empresa', 'Cargo', 'Setor', 'Status']
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    for row, emp in enumerate(employees, 2):
        ws.cell(row=row, column=1, value=emp.get('full_name', ''))
        ws.cell(row=row, column=2, value=emp.get('cpf', ''))
        ws.cell(row=row, column=3, value=emp.get('rg', ''))
        ws.cell(row=row, column=4, value=emp.get('registration_number', ''))
        ws.cell(row=row, column=5, value=companies.get(emp.get('company_id', ''), ''))
        ws.cell(row=row, column=6, value=emp.get('position', ''))
        ws.cell(row=row, column=7, value=emp.get('department', ''))
        ws.cell(row=row, column=8, value='Ativo' if emp.get('status') == 'active' else 'Inativo')
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=colaboradores_{datetime.now().strftime("%Y%m%d")}.xlsx'}
    )

@api_router.get('/employees/template/excel')
async def download_employees_template(current_user: dict = Depends(get_current_user)):
    """Download do template Excel para importação de colaboradores"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores"
    
    # Cabeçalho com instruções
    headers = ['Nome Completo*', 'CPF*', 'RG', 'Matrícula*', 'Empresa*', 'Cargo', 'Setor', 'Status']
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Exemplo de preenchimento
    ws.cell(row=2, column=1, value='João da Silva')
    ws.cell(row=2, column=2, value='123.456.789-00')
    ws.cell(row=2, column=3, value='12.345.678-9')
    ws.cell(row=2, column=4, value='MAT001')
    ws.cell(row=2, column=5, value='Nome da Empresa')
    ws.cell(row=2, column=6, value='Operador')
    ws.cell(row=2, column=7, value='Produção')
    ws.cell(row=2, column=8, value='Ativo')
    
    # Instruções
    ws2 = wb.create_sheet(title="Instruções")
    ws2.cell(row=1, column=1, value="INSTRUÇÕES DE PREENCHIMENTO").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="* Campos obrigatórios")
    ws2.cell(row=4, column=1, value="• Nome Completo: Nome completo do colaborador")
    ws2.cell(row=5, column=1, value="• CPF: Formato XXX.XXX.XXX-XX ou apenas números")
    ws2.cell(row=6, column=1, value="• Matrícula: Código único do colaborador na empresa")
    ws2.cell(row=7, column=1, value="• Empresa: Nome exato da empresa cadastrada no sistema")
    ws2.cell(row=8, column=1, value="• Status: 'Ativo' ou 'Inativo' (padrão: Ativo)")
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=template_colaboradores.xlsx'}
    )

@api_router.post('/employees/import/excel')
async def import_employees_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Importa colaboradores de um arquivo Excel"""
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Permissão insuficiente')
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail='Arquivo deve ser Excel (.xlsx ou .xls)')
    
    db = await get_db()
    
    # Carregar empresas
    companies = {}
    async for company in db.companies.find():
        companies[company['legal_name'].lower().strip()] = str(company['_id'])
    
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        
        results = {'imported': 0, 'errors': [], 'skipped': 0}
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:  # Linha vazia
                continue
            
            full_name = str(row[0]).strip() if row[0] else None
            cpf = str(row[1]).strip() if row[1] else None
            rg = str(row[2]).strip() if row[2] else None
            registration_number = str(row[3]).strip() if row[3] else None
            company_name = str(row[4]).strip().lower() if row[4] else None
            position = str(row[5]).strip() if len(row) > 5 and row[5] else None
            department = str(row[6]).strip() if len(row) > 6 and row[6] else None
            status_str = str(row[7]).strip().lower() if len(row) > 7 and row[7] else 'ativo'
            
            # Validações
            errors = []
            if not full_name:
                errors.append('Nome é obrigatório')
            if not cpf:
                errors.append('CPF é obrigatório')
            if not registration_number:
                errors.append('Matrícula é obrigatória')
            if not company_name:
                errors.append('Empresa é obrigatória')
            elif company_name not in companies:
                errors.append(f'Empresa "{row[4]}" não encontrada no sistema')
            
            if errors:
                results['errors'].append({'row': row_num, 'errors': errors, 'name': full_name})
                continue
            
            # Verificar se já existe
            existing = await db.employees.find_one({
                "$or": [
                    {"cpf": cpf},
                    {"registration_number": registration_number, "company_id": companies.get(company_name)}
                ]
            })
            
            if existing:
                results['skipped'] += 1
                results['errors'].append({
                    'row': row_num, 
                    'errors': ['Colaborador já existe (CPF ou Matrícula duplicada)'],
                    'name': full_name
                })
                continue
            
            # Inserir colaborador
            employee = {
                'full_name': full_name,
                'cpf': cpf,
                'rg': rg,
                'registration_number': registration_number,
                'company_id': companies.get(company_name),
                'position': position,
                'department': department,
                'status': 'active' if status_str == 'ativo' else 'inactive',
                'facial_consent': False,
                'created_at': datetime.now(timezone.utc),
                'updated_at': datetime.now(timezone.utc)
            }
            
            await db.employees.insert_one(employee)
            results['imported'] += 1
        
        return results
        
    except Exception as e:
        logger.error(f"Erro ao importar Excel: {str(e)}")
        raise HTTPException(status_code=400, detail=f'Erro ao processar arquivo: {str(e)}')

# ===================== IMPRESSÃO PDF =====================

@api_router.get('/reports/employees/pdf')
async def generate_employees_pdf(current_user: dict = Depends(get_current_user)):
    """Gera relatório PDF de colaboradores"""
    if not can_view_sensitive_data(current_user['role']):
        raise HTTPException(status_code=403, detail='Permissão insuficiente')
    
    db = await get_db()
    employees = await db.employees.find({"status": "active"}).to_list(5000)
    companies = {str(c['_id']): c['legal_name'] async for c in db.companies.find()}
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    
    elements = []
    
    # Título
    elements.append(Paragraph("Relatório de Colaboradores", title_style))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tabela
    data = [['Nome', 'CPF', 'Matrícula', 'Empresa', 'Cargo', 'Setor']]
    
    for emp in employees:
        data.append([
            emp.get('full_name', '')[:30],
            emp.get('cpf', ''),
            emp.get('registration_number', ''),
            companies.get(emp.get('company_id', ''), '')[:25],
            (emp.get('position', '') or '')[:20],
            (emp.get('department', '') or '')[:15]
        ])
    
    table = Table(data, colWidths=[120, 90, 70, 120, 100, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.063, 0.725, 0.506)),  # Emerald
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Total de colaboradores ativos: {len(employees)}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=colaboradores_{datetime.now().strftime("%Y%m%d")}.pdf'}
    )

@api_router.get('/reports/deliveries/pdf')
async def generate_deliveries_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Gera relatório PDF de entregas"""
    db = await get_db()
    
    query = {}
    if start_date:
        query['created_at'] = {'$gte': datetime.fromisoformat(start_date.replace('Z', '+00:00'))}
    if end_date:
        if 'created_at' not in query:
            query['created_at'] = {}
        query['created_at']['$lte'] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    
    deliveries = await db.deliveries.find(query).sort('created_at', -1).to_list(1000)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    
    elements = []
    
    # Título
    elements.append(Paragraph("Relatório de Entregas de EPIs", title_style))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tabela
    data = [['Data', 'Colaborador', 'Itens', 'Tipo', 'Verificação Facial']]
    
    for delivery in deliveries:
        items_str = ', '.join([
            f"{item.get('quantity', 1)}x {item.get('epi_name', item.get('kit_name', 'Item'))}"
            for item in delivery.get('items', [])
        ])[:50]
        
        facial_match = delivery.get('facial_match_score')
        facial_str = f"{int(facial_match * 100)}%" if facial_match else 'N/A'
        
        data.append([
            format_datetime_safe(delivery.get('created_at'), '%d/%m/%Y %H:%M'),
            delivery.get('employee_name', '')[:25],
            items_str,
            'Devolução' if delivery.get('is_return') else 'Entrega',
            facial_str
        ])
    
    table = Table(data, colWidths=[100, 150, 200, 70, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.063, 0.725, 0.506)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Total de entregas: {len(deliveries)}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=entregas_{datetime.now().strftime("%Y%m%d")}.pdf'}
    )

@api_router.get('/reports/employee/{employee_id}/pdf')
async def generate_employee_history_pdf(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Gera PDF com ficha do colaborador e histórico de entregas"""
    db = await get_db()
    
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    
    company = None
    if employee.get('company_id'):
        company = await db.companies.find_one({"_id": ObjectId(employee['company_id'])})
    
    deliveries = await db.deliveries.find({"employee_id": employee_id}).sort('created_at', -1).to_list(500)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=40)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, spaceAfter=10)
    
    elements = []
    
    # Cabeçalho
    elements.append(Paragraph("Ficha do Colaborador", title_style))
    elements.append(Spacer(1, 10))
    
    # Dados do colaborador
    info_data = [
        ['Nome:', employee.get('full_name', '')],
        ['CPF:', employee.get('cpf', '')],
        ['RG:', employee.get('rg', '') or '-'],
        ['Matrícula:', employee.get('registration_number', '')],
        ['Empresa:', company.get('legal_name', '') if company else '-'],
        ['Cargo:', employee.get('position', '') or '-'],
        ['Setor:', employee.get('department', '') or '-'],
        ['Status:', 'Ativo' if employee.get('status') == 'active' else 'Inativo'],
    ]
    
    info_table = Table(info_data, colWidths=[100, 350])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 30))
    
    # Histórico de entregas
    elements.append(Paragraph("Histórico de Entregas/Devoluções", subtitle_style))
    
    if deliveries:
        # ATUALIZADO: Incluir responsável pela entrega
        history_data = [['Data', 'Tipo', 'Itens', 'Responsável', 'Verificação']]
        
        for delivery in deliveries:
            items_str = ', '.join([
                f"{item.get('quantity', 1)}x {item.get('epi_name', item.get('kit_name', 'Item'))}"
                for item in delivery.get('items', [])
            ])[:50]
            
            facial_match = delivery.get('facial_match_score')
            facial_str = f"{int(facial_match * 100)}%" if facial_match else '-'
            
            # Nome do responsável pela entrega
            responsavel = delivery.get('delivered_by_name', '-')
            
            history_data.append([
                format_datetime_safe(delivery.get('created_at'), '%d/%m/%Y'),
                'Devolução' if delivery.get('is_return') else 'Entrega',
                items_str,
                responsavel[:20],  # Limitar tamanho do nome
                facial_str
            ])
        
        history_table = Table(history_data, colWidths=[65, 60, 180, 100, 55])
        history_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.063, 0.725, 0.506)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
        ]))
        
        elements.append(history_table)
    else:
        elements.append(Paragraph("Nenhuma entrega registrada.", styles['Normal']))
    
    elements.append(Spacer(1, 40))
    elements.append(Paragraph(f"Documento gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=ficha_{employee.get("registration_number", employee_id)}_{datetime.now().strftime("%Y%m%d")}.pdf'}
    )

# ===================== FACIAL TEMPLATES =====================

@api_router.get('/employees/{employee_id}/facial-templates')
async def get_facial_templates(employee_id: str, current_user: dict = Depends(get_current_user)):
    db = await get_db()
    templates = await db.facial_templates.find({
        "employee_id": {"$in": [employee_id, ObjectId(employee_id)]}
    }).to_list(100)
    return [format_facial_template_response(t) for t in templates]

@api_router.get('/employees/{employee_id}/biometric-summary')
async def get_employee_biometric_summary(employee_id: str, current_user: dict = Depends(get_current_user)):
    db = await get_db()
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador nao encontrado')
    templates = await db.facial_templates.find({
        "employee_id": {"$in": [employee_id, ObjectId(employee_id)]}
    }).to_list(100)
    return build_biometric_summary(employee, templates)

@api_router.delete('/employees/{employee_id}/facial-templates')
async def delete_all_facial_templates(employee_id: str, current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissao')
    db = await get_db()
    result = await db.facial_templates.delete_many({
        "employee_id": {"$in": [employee_id, ObjectId(employee_id)]}
    })
    await face_recognition_service.refresh_employee_cache(db, employee_id)
    return {'message': 'Biometria facial excluida', 'deleted_count': result.deleted_count}

@api_router.post('/employees/{employee_id}/facial-templates/reprocess')
async def reprocess_facial_templates(employee_id: str, current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissao')
    db = await get_db()
    await face_recognition_service.refresh_employee_cache(db, employee_id)
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    templates = await db.facial_templates.find({
        "employee_id": {"$in": [employee_id, ObjectId(employee_id)]}
    }).to_list(100)
    return {
        'message': 'Templates reprocessados no cache facial',
        'summary': build_biometric_summary(employee, templates) if employee else None
    }

@api_router.get('/biometric/dashboard')
async def get_biometric_dashboard(current_user: dict = Depends(get_current_user)):
    db = await get_db()
    employees = await db.employees.find({"status": "active"}).to_list(5000)
    summaries = await get_biometric_summaries(db, employees)
    total = len(employees)
    registered = sum(1 for item in summaries.values() if item["status"] == "registered")
    incomplete = sum(1 for item in summaries.values() if item["status"] == "incomplete")
    missing = total - registered - incomplete
    coverage = round((registered / total) * 100, 1) if total else 0
    return {
        "total_employees": total,
        "registered": registered,
        "missing": missing,
        "incomplete": incomplete,
        "coverage_percent": coverage
    }

# Endpoint otimizado para buscar TODOS os templates de uma vez
@api_router.get('/facial-templates/all')
async def get_all_facial_templates(current_user: dict = Depends(get_current_user)):
    """Retorna todos os templates faciais com informações do colaborador - otimizado para reconhecimento"""
    db = await get_db()
    
    # Buscar todos os templates
    templates = await db.facial_templates.find({}).to_list(1000)
    
    if not templates:
        return []
    
    # Buscar todos os colaboradores de uma vez
    employee_ids = []
    for t in templates:
        emp_id = t.get('employee_id')
        if emp_id:
            # Converter para ObjectId se for string
            if isinstance(emp_id, str):
                employee_ids.append(ObjectId(emp_id))
            else:
                employee_ids.append(emp_id)
    
    employee_ids = list(set(employee_ids))
    employees = await db.employees.find({
        "_id": {"$in": employee_ids}
    }).to_list(1000)
    
    # Criar mapa de colaboradores
    emp_map = {str(e['_id']): doc_to_response(e) for e in employees}
    
    # Montar resposta com dados do colaborador
    result = []
    for t in templates:
        emp_id = t.get('employee_id')
        emp_id_str = str(emp_id) if emp_id else None
        if emp_id_str and emp_id_str in emp_map:
            result.append({
                'id': str(t['_id']),
                'employee_id': emp_id_str,
                'descriptor': t.get('descriptor'),
                'pose_label': t.get('pose_label'),
                'quality_score': t.get('quality_score'),
                'detection_score': t.get('detection_score'),
                'employee': emp_map[emp_id_str]
            })
    
    return result

# ===================== VERIFICAÇÃO DE DUPLICIDADE BIOMÉTRICA =====================

BIOMETRIC_DUPLICATE_THRESHOLD = 0.40  # Distância máxima para considerar duplicata

def euclidean_distance(a, b):
    """Calcula distância euclidiana entre dois vetores"""
    import math
    return math.sqrt(sum((x - y) ** 2 for x, y in zip(a, b)))

@api_router.post('/biometric/check-duplicate')
async def check_biometric_duplicate(
    data: BiometricCheckRequest,
    current_user: dict = Depends(get_current_user)
):
    """Verifica se existe biometria duplicada no sistema"""
    db = await get_db()
    
    try:
        # Parsear o descriptor enviado
        new_descriptor = json.loads(data.descriptor)
        if len(new_descriptor) != 128:
            raise HTTPException(status_code=400, detail="Descriptor deve ter 128 valores")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Descriptor inválido")
    
    # Buscar todos os templates existentes
    templates = await db.facial_templates.find({}).to_list(1000)
    
    # Buscar colaboradores
    employee_ids = list(set([t.get('employee_id') for t in templates if t.get('employee_id')]))
    employees = await db.employees.find({"_id": {"$in": [ObjectId(str(e)) for e in employee_ids]}}).to_list(1000)
    emp_map = {str(e['_id']): e for e in employees}
    
    # Comparar com cada template
    for template in templates:
        emp_id = str(template.get('employee_id'))
        
        # Ignorar o próprio colaborador (para edição)
        if data.employee_id and emp_id == data.employee_id:
            continue
        
        try:
            existing_descriptor = json.loads(template.get('descriptor', '[]'))
            if len(existing_descriptor) != 128:
                continue
            
            # Calcular distância
            distance = euclidean_distance(new_descriptor, existing_descriptor)
            similarity = 1 - distance
            
            # Se a distância for menor que o threshold, é duplicata
            if distance < BIOMETRIC_DUPLICATE_THRESHOLD:
                emp_data = emp_map.get(emp_id, {})
                return BiometricCheckResponse(
                    is_duplicate=True,
                    duplicate_employee_id=emp_id,
                    duplicate_employee_name=emp_data.get('full_name', 'Desconhecido'),
                    similarity_score=similarity,
                    message=f"Esta foto já corresponde ou está muito semelhante à biometria facial do colaborador '{emp_data.get('full_name', 'Desconhecido')}'. Não é permitido utilizar a mesma foto para colaboradores diferentes."
                )
        except Exception as e:
            continue
    
    return BiometricCheckResponse(
        is_duplicate=False,
        message="Biometria válida - nenhuma duplicata encontrada"
    )

@api_router.get('/biometric/audit-duplicates')
async def audit_biometric_duplicates(current_user: dict = Depends(get_current_user)):
    """Audita a base de dados para encontrar biometrias duplicadas existentes"""
    if current_user['role'] not in ['admin', 'gestor']:
        raise HTTPException(status_code=403, detail="Apenas administradores podem auditar duplicatas")
    
    db = await get_db()
    
    # Buscar todos os templates
    templates = await db.facial_templates.find({}).to_list(1000)
    
    # Buscar colaboradores
    employee_ids = list(set([t.get('employee_id') for t in templates if t.get('employee_id')]))
    employees = await db.employees.find({"_id": {"$in": [ObjectId(str(e)) for e in employee_ids]}}).to_list(1000)
    emp_map = {str(e['_id']): e for e in employees}
    
    duplicates = []
    checked_pairs = set()
    
    # Comparar todos os pares
    for i, t1 in enumerate(templates):
        emp_id1 = str(t1.get('employee_id'))
        
        try:
            desc1 = json.loads(t1.get('descriptor', '[]'))
            if len(desc1) != 128:
                continue
        except:
            continue
        
        for j, t2 in enumerate(templates):
            if i >= j:
                continue
            
            emp_id2 = str(t2.get('employee_id'))
            
            # Não comparar mesmo colaborador
            if emp_id1 == emp_id2:
                continue
            
            # Evitar duplicar pares
            pair_key = tuple(sorted([emp_id1, emp_id2]))
            if pair_key in checked_pairs:
                continue
            checked_pairs.add(pair_key)
            
            try:
                desc2 = json.loads(t2.get('descriptor', '[]'))
                if len(desc2) != 128:
                    continue
                
                distance = euclidean_distance(desc1, desc2)
                
                if distance < BIOMETRIC_DUPLICATE_THRESHOLD:
                    emp1 = emp_map.get(emp_id1, {})
                    emp2 = emp_map.get(emp_id2, {})
                    
                    duplicates.append({
                        "employee1_id": emp_id1,
                        "employee1_name": emp1.get('full_name', 'Desconhecido'),
                        "employee2_id": emp_id2,
                        "employee2_name": emp2.get('full_name', 'Desconhecido'),
                        "similarity_score": round(1 - distance, 4),
                        "distance": round(distance, 4)
                    })
            except:
                continue
    
    return {
        "total_templates": len(templates),
        "duplicates_found": len(duplicates),
        "duplicates": duplicates
    }

@api_router.post('/employees/{employee_id}/biometric-consent')
async def register_biometric_consent(
    employee_id: str,
    consent_data: BiometricConsentRequest,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Registra o consentimento biométrico do colaborador (LGPD)"""
    db = await get_db()
    
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    # Obter IP do cliente
    client_ip = consent_data.ip_address or request.client.host
    
    # Atualizar colaborador com dados de consentimento
    update_data = {
        "facial_consent": consent_data.accepted,
        "facial_consent_date": datetime.now(timezone.utc),
        "facial_consent_ip": client_ip
    }
    
    await db.employees.update_one(
        {"_id": ObjectId(employee_id)},
        {"$set": update_data}
    )
    
    # Registrar log de consentimento para auditoria
    consent_log = {
        "employee_id": ObjectId(employee_id),
        "employee_name": employee.get('full_name'),
        "accepted": consent_data.accepted,
        "ip_address": client_ip,
        "user_agent": request.headers.get("user-agent", "unknown"),
        "registered_by": current_user.get("username"),
        "created_at": datetime.now(timezone.utc)
    }
    await db.biometric_consent_logs.insert_one(consent_log)
    
    return {
        "success": True,
        "message": "Consentimento biométrico registrado com sucesso",
        "consent_date": update_data["facial_consent_date"].isoformat(),
        "ip_address": client_ip
    }

@api_router.get('/employees/{employee_id}/biometric-consent')
async def get_biometric_consent(
    employee_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Verifica o status do consentimento biométrico do colaborador"""
    db = await get_db()
    
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    return {
        "has_consent": employee.get("facial_consent", False),
        "consent_date": employee.get("facial_consent_date"),
        "consent_ip": employee.get("facial_consent_ip")
    }

@api_router.post('/employees/{employee_id}/facial-templates')
async def create_facial_template(employee_id: str, template_data: FacialTemplateCreate, current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão')
    
    db = await get_db()
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    
    new_template = {
        "employee_id": ObjectId(employee_id),
        "descriptor": template_data.descriptor,
        "pose_label": template_data.pose_label,
        "quality_score": template_data.quality_score,
        "detection_score": template_data.detection_score,
        "created_by": current_user.get("id"),
        "created_by_name": current_user.get("username"),
        "created_at": datetime.now(timezone.utc)
    }
    try:
        result = await db.facial_templates.insert_one(new_template)
        await face_recognition_service.refresh_employee_cache(db, employee_id)
    except Exception as error:
        logger.exception("Falha ao salvar template facial do colaborador %s: %s", employee_id, error)
        raise HTTPException(status_code=500, detail='Erro ao salvar template facial')
    
    # Retornar resposta formatada corretamente
    return {
        "id": str(result.inserted_id),
        "employee_id": employee_id,
        "descriptor": template_data.descriptor,
        "pose_label": template_data.pose_label,
        "quality_score": template_data.quality_score,
        "detection_score": template_data.detection_score,
        "created_by": new_template.get("created_by"),
        "created_by_name": new_template.get("created_by_name"),
        "created_at": new_template["created_at"].isoformat()
    }

@api_router.delete('/employees/{employee_id}/facial-templates/{template_id}')
async def delete_facial_template(employee_id: str, template_id: str, current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão')
    
    db = await get_db()
    result = await db.facial_templates.delete_one({
        "_id": ObjectId(template_id),
        "employee_id": {"$in": [employee_id, ObjectId(employee_id)]}
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Template não encontrado')
    
    await face_recognition_service.refresh_employee_cache(db, employee_id)
    return {'message': 'Template excluído'}

@api_router.post('/facial/enroll', response_model=FacialEnrollResponse, status_code=status.HTTP_201_CREATED)
async def facial_enroll(
    payload: FacialEnrollRequest,
    current_user: dict = Depends(get_current_user)
):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissao')

    db = await get_db()
    employee = await db.employees.find_one({"_id": ObjectId(payload.employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador nao encontrado')

    if not face_recognition_service.available:
        raise HTTPException(status_code=503, detail='Servico facial backend indisponivel')

    try:
        enroll_result = face_recognition_service.enroll_from_image(
            payload.image_base64,
            pose_label=payload.pose_label or 'frontal'
        )
        if enroll_result.get('status') != 'ok':
            logger.warning("Falha no enroll facial do colaborador %s: %s", payload.employee_id, enroll_result.get('message'))
            raise HTTPException(status_code=400, detail=enroll_result.get('message', 'Falha ao cadastrar biometria facial'))
    except HTTPException:
        raise
    except Exception as error:
        logger.exception("Erro ao gerar embedding facial do colaborador %s: %s", payload.employee_id, error)
        raise HTTPException(status_code=500, detail='Erro ao processar biometria facial')

    descriptor_json = json.dumps(enroll_result['embedding'].tolist())
    new_template = {
        "employee_id": ObjectId(payload.employee_id),
        "descriptor": descriptor_json,
        "pose_label": enroll_result.get('pose_label') or payload.pose_label or 'frontal',
        "quality_score": enroll_result.get('quality_score'),
        "detection_score": enroll_result.get('detection_score'),
        "created_by": current_user.get("id"),
        "created_by_name": current_user.get("username"),
        "created_at": datetime.now(timezone.utc)
    }
    try:
        result = await db.facial_templates.insert_one(new_template)
        await face_recognition_service.refresh_employee_cache(db, payload.employee_id)
    except Exception as error:
        logger.exception("Falha ao gravar enroll facial do colaborador %s: %s", payload.employee_id, error)
        raise HTTPException(status_code=500, detail='Erro ao salvar template facial')

    return FacialEnrollResponse(
        status='ok',
        message=enroll_result.get('message', 'Biometria facial cadastrada com sucesso.'),
        employee_id=payload.employee_id,
        template_id=str(result.inserted_id),
        pose_label=new_template['pose_label'],
        quality_score=new_template['quality_score'],
        detection_score=new_template['detection_score']
    )

@api_router.post('/facial/identify-fast', response_model=FacialIdentifyFastResponse)
async def facial_identify_fast(
    payload: FacialIdentifyFastRequest,
    current_user: dict = Depends(get_current_user)
):
    if not can_deliver_epi(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissao para reconhecimento facial de entrega')

    if not face_recognition_service.available:
        raise HTTPException(status_code=503, detail='Servico facial backend indisponivel')

    result = face_recognition_service.identify_fast(payload.image_base64)
    return FacialIdentifyFastResponse(**result)

@api_router.post('/facial/identify-burst', response_model=FacialIdentifyFastResponse)
async def facial_identify_burst(
    payload: FacialIdentifyBurstRequest,
    current_user: dict = Depends(get_current_user)
):
    if not can_deliver_epi(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissao para reconhecimento facial de entrega')

    if not face_recognition_service.available:
        raise HTTPException(status_code=503, detail='Servico facial backend indisponivel')

    frames = [image for image in payload.images_base64[:5] if image]
    if not frames:
        return FacialIdentifyFastResponse(
            status='retry',
            message='Nenhuma imagem utilizavel recebida. Nova captura automatica em andamento.',
            employee_id=None,
            employee_name=None,
            similarity_score=0.0,
            detection_confidence=0.0,
            liveness_required=False
        )

    results = [face_recognition_service.identify_fast(image) for image in frames]
    approved = [item for item in results if item.get('status') == 'approved' and item.get('employee_id')]
    if approved:
        best = max(approved, key=lambda item: item.get('similarity_score', 0))
        best['message'] = f"{best.get('employee_name')} identificado automaticamente."
        best['liveness_required'] = False
        return FacialIdentifyFastResponse(**best)

    grouped = {}
    for item in results:
        employee_id = item.get('employee_id')
        if not employee_id:
            continue
        grouped.setdefault(employee_id, []).append(item)

    for employee_results in grouped.values():
        if len(employee_results) < 2:
            continue
        scores = [item.get('similarity_score', 0) for item in employee_results]
        average_score = sum(scores) / len(scores)
        best = max(employee_results, key=lambda item: item.get('similarity_score', 0))
        if average_score >= 0.70 and max(scores) >= 0.74:
            best.update({
                'status': 'approved',
                'similarity_score': round(float(average_score), 4),
                'message': f"{best.get('employee_name')} confirmado por multiplos frames.",
                'liveness_required': False
            })
            return FacialIdentifyFastResponse(**best)

    usable = [item for item in results if item.get('detection_confidence', 0) > 0]
    best_retry = max(usable, key=lambda item: item.get('similarity_score', 0), default=None)
    if best_retry:
        best_retry.update({
            'status': 'retry',
            'employee_id': None if best_retry.get('similarity_score', 0) < 0.70 else best_retry.get('employee_id'),
            'employee_name': None if best_retry.get('similarity_score', 0) < 0.70 else best_retry.get('employee_name'),
            'message': 'Procurando colaborador automaticamente.',
            'liveness_required': False
        })
        return FacialIdentifyFastResponse(**best_retry)

    return FacialIdentifyFastResponse(
        status='retry',
        message='Nenhum frame utilizavel. Nova captura automatica em andamento.',
        employee_id=None,
        employee_name=None,
        similarity_score=0.0,
        detection_confidence=0.0,
        liveness_required=False
    )

@api_router.post('/facial/liveness-check', response_model=FacialLivenessResponse)
async def facial_liveness_check(
    payload: FacialLivenessRequest,
    current_user: dict = Depends(get_current_user)
):
    if not can_deliver_epi(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissao para prova de vida')

    if not face_recognition_service.available:
        raise HTTPException(status_code=503, detail='Servico facial backend indisponivel')

    result = face_recognition_service.liveness_check(
        payload.image_base64,
        previous_image_base64=payload.previous_image_base64
    )
    return FacialLivenessResponse(**result)

@api_router.get('/facial/migration-status', response_model=FacialMigrationStatusResponse)
async def facial_migration_status(current_user: dict = Depends(require_role('admin', 'gestor'))):
    db = await get_db()
    status_data = await face_recognition_service.get_migration_status(db)
    return FacialMigrationStatusResponse(**status_data)

@api_router.post('/facial/reload-cache')
async def facial_reload_cache(current_user: dict = Depends(require_role('admin', 'gestor'))):
    db = await get_db()
    await face_recognition_service.reload_cache(db)
    return {
        'message': 'Cache facial recarregado com sucesso',
        'service_available': face_recognition_service.available,
        'cache_size': face_recognition_service.cache_size,
        'employees_loaded': face_recognition_service.cached_employee_count,
    }

# ===================== SUPPLIERS =====================

@api_router.get('/suppliers', response_model=List[SupplierResponse])
async def get_suppliers(current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    suppliers = await db.suppliers.find({}).to_list(1000)
    return [SupplierResponse(**doc_to_response(s)) for s in suppliers]

@api_router.post('/suppliers', response_model=SupplierResponse, status_code=status.HTTP_201_CREATED)
async def create_supplier(supplier_data: SupplierCreate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    new_supplier = {**supplier_data.model_dump(), "created_at": datetime.now(timezone.utc)}
    result = await db.suppliers.insert_one(new_supplier)
    new_supplier['_id'] = result.inserted_id
    return SupplierResponse(**doc_to_response(new_supplier))

@api_router.patch('/suppliers/{supplier_id}', response_model=SupplierResponse)
async def update_supplier(supplier_id: str, supplier_data: SupplierUpdate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    update_data = {k: v for k, v in supplier_data.model_dump(exclude_unset=True).items()}
    update_data['updated_at'] = datetime.now(timezone.utc)
    result = await db.suppliers.find_one_and_update(
        {"_id": ObjectId(supplier_id)}, {"$set": update_data}, return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail='Fornecedor não encontrado')
    return SupplierResponse(**doc_to_response(result))

@api_router.delete('/suppliers/{supplier_id}')
async def delete_supplier(supplier_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    result = await db.suppliers.delete_one({"_id": ObjectId(supplier_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Fornecedor não encontrado')
    return {'message': 'Fornecedor excluído'}

# ===================== EPIS =====================

def normalize_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None

def normalize_group_text(value: Optional[str]) -> str:
    text = normalize_text(value) or ''
    return ' '.join(text.lower().split())

def normalize_size_value(value: Optional[str]) -> Optional[str]:
    text = normalize_text(value)
    return text.upper() if text else None

def strip_size_from_epi_name(name: Optional[str], size: Optional[str]) -> str:
    base = normalize_text(name) or ''
    normalized_size = normalize_text(size)
    if not base or not normalized_size:
        return base

    candidates = [
        f" tamanho {normalized_size}",
        f" tam {normalized_size}",
        f" tam. {normalized_size}",
        f" size {normalized_size}",
    ]
    lowered = base.lower()
    for suffix in candidates:
        if lowered.endswith(suffix.lower()):
            return base[:len(base) - len(suffix)].strip()

    parts = base.split()
    if parts and parts[-1].upper() == normalized_size.upper():
        return ' '.join(parts[:-1]).strip() or base
    return base

def get_epi_group_name(epi: dict) -> str:
    return normalize_text(epi.get('epi_group_name')) or strip_size_from_epi_name(epi.get('name') or epi.get('description'), epi.get('size') or epi.get('tamanho'))

def get_epi_group_key(epi: dict) -> str:
    explicit_key = normalize_text(epi.get('epi_group_key'))
    if explicit_key:
        return explicit_key
    parts = [
        get_epi_group_name(epi),
        infer_epi_category(epi),
        epi.get('model') or epi.get('modelo'),
        epi.get('brand') or epi.get('marca'),
    ]
    return '|'.join(normalize_group_text(part) for part in parts if normalize_text(part))

def slugify_group_key(value: str) -> str:
    normalized = unicodedata.normalize('NFKD', value or '')
    ascii_text = normalized.encode('ascii', 'ignore').decode('ascii')
    slug = re.sub(r'[^A-Za-z0-9]+', '-', ascii_text.upper()).strip('-')
    return slug or 'EPI'

async def generate_unique_epi_group_key(db, group_name: str) -> str:
    base = slugify_group_key(group_name)
    existing_keys = {
        epi.get('epi_group_key')
        for epi in await db.epis.find({"epi_group_key": {"$regex": f"^{base}-", "$options": "i"}}).to_list(1000)
        if epi.get('epi_group_key')
    }
    index = 1
    while True:
        candidate = f"{base}-{index:03d}"
        if candidate not in existing_keys:
            return candidate
        index += 1

async def resolve_epi_group_fields(db, payload: dict, fallback_name: Optional[str] = None) -> tuple[Optional[str], Optional[str]]:
    group_key = normalize_text(payload.get('epi_group_key'))
    group_name = normalize_text(payload.get('epi_group_name'))
    if group_key:
        if not group_name:
            existing = await db.epis.find_one({"epi_group_key": group_key})
            group_name = existing.get('epi_group_name') if existing else fallback_name
        return group_key, group_name
    if group_name:
        return await generate_unique_epi_group_key(db, group_name), group_name
    return None, None

def infer_epi_category(epi: dict) -> str:
    return epi.get('category') or epi.get('type_category') or 'Sem categoria'

def infer_epi_nbr(epi: dict) -> Optional[str]:
    return epi.get('nbr') or epi.get('nbr_number')

def infer_epi_has_size(epi: dict, variations: List[dict]) -> bool:
    if epi.get('possui_variacao_tamanho') is not None:
        return bool(epi.get('possui_variacao_tamanho'))
    if normalize_text(epi.get('size')):
        return True
    return any(normalize_text(v.get('size')) for v in variations)

async def get_supplier_name(db, supplier_id: Optional[str]) -> Optional[str]:
    if not supplier_id:
        return None
    supplier = await db.suppliers.find_one({"_id": ObjectId(supplier_id)})
    return supplier.get('name') if supplier else None

async def create_variation_from_legacy_epi(db, epi: dict) -> Optional[dict]:
    epi_id = str(epi['_id'])
    existing = await db.epi_variations.find_one({"epi_id": epi_id})
    if existing:
        return existing
    has_legacy_data = any([
        normalize_text(epi.get('brand')),
        normalize_text(epi.get('model')),
        normalize_text(epi.get('ca_number')),
        normalize_text(epi.get('supplier_id')),
        normalize_text(epi.get('batch')),
        normalize_text(epi.get('size')),
        epi.get('current_stock') is not None
    ])
    if not has_legacy_data:
        return None

    variation = {
        "epi_id": epi_id,
        "brand": epi.get('brand'),
        "model": epi.get('model'),
        "ca_number": epi.get('ca_number'),
        "supplier_id": epi.get('supplier_id'),
        "supplier_name": await get_supplier_name(db, epi.get('supplier_id')),
        "ca_validity": epi.get('ca_validity'),
        "size": epi.get('size'),
        "color": epi.get('color'),
        "current_stock": epi.get('current_stock', 0) or 0,
        "unit_price": epi.get('unit_price'),
        "purchase_date": epi.get('purchase_date'),
        "batch": epi.get('batch'),
        "status": epi.get('status') or 'ativo',
        "qr_code": epi.get('qr_code'),
        "internal_code": epi.get('internal_code'),
        "invoice_number": epi.get('invoice_number'),
        "validity_date": epi.get('validity_date'),
        "material": epi.get('material'),
        "technical_standard": epi.get('technical_standard'),
        "quantity_purchased": epi.get('quantity_purchased', epi.get('current_stock', 0) or 0),
        "created_at": epi.get('created_at') or datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    result = await db.epi_variations.insert_one(variation)
    variation['_id'] = result.inserted_id
    return variation

async def get_epi_variations(db, epi: dict) -> List[dict]:
    await create_variation_from_legacy_epi(db, epi)
    return await db.epi_variations.find({"epi_id": str(epi['_id'])}).to_list(200)

async def get_epis_by_group_key(db, group_key: Optional[str]) -> List[dict]:
    if not group_key:
        return []
    epis = await db.epis.find({}).to_list(5000)
    return [epi for epi in epis if get_epi_group_key(epi) == group_key]

async def build_epi_group_summary(db, group_key: str, fallback_item: Optional[dict] = None) -> Optional[dict]:
    fallback_item = fallback_item or {}
    group_epis = await get_epis_by_group_key(db, group_key)
    seed = group_epis[0] if group_epis else (fallback_item or {})
    if not seed:
        return None

    available_count = 0
    variations_count = 0
    for epi in group_epis:
        variations = await get_epi_variations(db, epi)
        variations_count += len(variations)
        available_count += len([
            variation for variation in variations
            if (variation.get('current_stock', 0) or 0) > 0 and (variation.get('status') or 'ativo') != 'inativo'
        ])

    return {
        "epi_group_key": group_key,
        "epi_base_key": group_key,
        "name": fallback_item.get('name') or get_epi_group_name(seed),
        "epi_group_name": fallback_item.get('epi_group_name') or get_epi_group_name(seed),
        "description": fallback_item.get('description') or seed.get('description'),
        "category": fallback_item.get('category') or infer_epi_category(seed),
        "type_category": fallback_item.get('type_category') or infer_epi_category(seed),
        "model": fallback_item.get('model') or seed.get('model'),
        "modelo": fallback_item.get('model') or seed.get('model'),
        "brand": fallback_item.get('brand') or seed.get('brand'),
        "marca": fallback_item.get('brand') or seed.get('brand'),
        "variations_count": variations_count,
        "available_variations_count": available_count
    }

def serialize_epi_variation(variation: dict) -> dict:
    return {
        "id": str(variation['_id']),
        "epi_id": variation.get('epi_id'),
        "brand": variation.get('brand'),
        "marca": variation.get('brand'),
        "model": variation.get('model'),
        "modelo": variation.get('model'),
        "ca_number": variation.get('ca_number'),
        "ca": variation.get('ca_number'),
        "supplier_id": variation.get('supplier_id'),
        "supplier_name": variation.get('supplier_name'),
        "fornecedor": variation.get('supplier_name'),
        "ca_validity": variation.get('ca_validity'),
        "validade_ca": variation.get('ca_validity'),
        "size": variation.get('size'),
        "tamanho": variation.get('size'),
        "color": variation.get('color'),
        "cor": variation.get('color'),
        "current_stock": variation.get('current_stock', 0) or 0,
        "quantidade_estoque": variation.get('current_stock', 0) or 0,
        "unit_price": variation.get('unit_price'),
        "valor_unitario": variation.get('unit_price'),
        "purchase_date": variation.get('purchase_date'),
        "data_compra": variation.get('purchase_date'),
        "batch": variation.get('batch'),
        "lote": variation.get('batch'),
        "status": variation.get('status'),
        "qr_code": variation.get('qr_code'),
        "internal_code": variation.get('internal_code'),
        "invoice_number": variation.get('invoice_number'),
        "validity_date": variation.get('validity_date'),
        "material": variation.get('material'),
        "technical_standard": variation.get('technical_standard'),
        "quantity_purchased": variation.get('quantity_purchased', 0) or 0
    }

def choose_primary_variation(variations: List[dict]) -> Optional[dict]:
    if not variations:
        return None
    ranked = sorted(
        variations,
        key=lambda item: (
            0 if (item.get('status') or 'ativo') == 'ativo' else 1,
            -(item.get('current_stock', 0) or 0),
            normalize_text(item.get('size')) is None
        )
    )
    return ranked[0]

async def sync_epi_stock_summary(db, epi_id: str) -> None:
    variations = await db.epi_variations.find({"epi_id": epi_id}).to_list(200)
    total_stock = sum((item.get('current_stock', 0) or 0) for item in variations)
    total_purchased = sum((item.get('quantity_purchased', 0) or 0) for item in variations)
    await db.epis.update_one(
        {"_id": ObjectId(epi_id)},
        {"$set": {
            "current_stock": total_stock,
            "quantity_purchased": total_purchased,
            "updated_at": datetime.now(timezone.utc)
        }}
    )

async def build_epi_response(db, epi: dict) -> dict:
    variations = await get_epi_variations(db, epi)
    for variation in variations:
        if variation.get('supplier_id') and not variation.get('supplier_name'):
            variation['supplier_name'] = await get_supplier_name(db, variation.get('supplier_id'))
    primary = choose_primary_variation(variations)
    serialized_variations = [serialize_epi_variation(item) for item in variations]
    total_stock = sum(item.get('current_stock', 0) or 0 for item in variations)
    total_purchased = sum(item.get('quantity_purchased', 0) or 0 for item in variations)
    validity_candidates = [
        parsed_date
        for item in variations
        for parsed_date in [ensure_aware_datetime(item.get('validity_date') or item.get('ca_validity'))]
        if parsed_date
    ]
    validity_date = min(validity_candidates) if validity_candidates else None
    response = doc_to_response(epi)
    response.update({
        "category": infer_epi_category(epi),
        "type_category": infer_epi_category(epi),
        "description": epi.get('description'),
        "obrigatorio_ca": epi.get('obrigatorio_ca', True),
        "nbr": infer_epi_nbr(epi),
        "nbr_number": infer_epi_nbr(epi),
        "possui_variacao_tamanho": infer_epi_has_size(epi, variations),
        "epi_group_key": epi.get('epi_group_key'),
        "epi_group_name": epi.get('epi_group_name'),
        "variations": serialized_variations,
        "current_stock": total_stock,
        "quantity_purchased": total_purchased,
        "min_stock": epi.get('min_stock', 0) or 0,
        "max_stock": epi.get('max_stock')
    })
    if primary:
        response.update({
            "brand": primary.get('brand'),
            "model": primary.get('model'),
            "color": primary.get('color'),
            "size": primary.get('size'),
            "material": primary.get('material'),
            "ca_number": primary.get('ca_number'),
            "ca_validity": primary.get('ca_validity'),
            "technical_standard": primary.get('technical_standard'),
            "supplier_id": primary.get('supplier_id'),
            "supplier_name": primary.get('supplier_name'),
            "invoice_number": primary.get('invoice_number'),
            "purchase_date": primary.get('purchase_date'),
            "unit_price": primary.get('unit_price'),
            "total_price": (primary.get('unit_price') or 0) * (primary.get('quantity_purchased') or 0),
            "internal_code": primary.get('internal_code'),
            "batch": primary.get('batch'),
            "qr_code": primary.get('qr_code'),
            "validity_date": validity_date or primary.get('validity_date')
        })
    else:
        response["validity_date"] = validity_date
    stock_status, validity_status = calculate_epi_status(response)
    response['stock_status'] = stock_status
    response['validity_status'] = validity_status
    return response

async def build_kit_response(db, kit: dict) -> dict:
    response = doc_to_response(kit)
    items = []
    grouped_items = {}
    for raw_item in kit.get('items', []):
        group_key = raw_item.get('epi_group_key') or raw_item.get('epi_base_key')
        epi_base_id = raw_item.get('epi_base_id') or raw_item.get('epi_id')
        if not group_key and epi_base_id and ObjectId.is_valid(str(epi_base_id)):
            epi = await db.epis.find_one({"_id": ObjectId(epi_base_id)})
            if epi:
                group_key = get_epi_group_key(epi)

        if not group_key:
            continue
        if group_key in grouped_items:
            grouped_items[group_key]["quantity"] = max(grouped_items[group_key]["quantity"], raw_item.get('quantity', 1))
            continue

        summary = await build_epi_group_summary(db, group_key, raw_item)
        if not summary:
            continue
        item = {
            "epi_group_key": group_key,
            "epi_base_key": group_key,
            "epi_base_id": None,
            "epi_id": None,
            "quantity": raw_item.get('quantity', 1),
            **summary
        }
        grouped_items[group_key] = item
        items.append(item)
    response['items'] = items
    return response

def infer_employee_size_field(epi: dict) -> Optional[str]:
    category = (infer_epi_category(epi) or '').lower()
    name = (epi.get('name') or '').lower()
    if any(token in name for token in ['botina', 'bota', 'sapato', 'calçado', 'calcado']) or 'pés' in category or 'pes' in category:
        return 'tamanho_calcado'
    if 'luva' in name or 'mãos' in category or 'maos' in category:
        return 'tamanho_luva'
    if any(token in name for token in ['calça', 'calca', 'bermuda']) or 'pernas' in category:
        return 'tamanho_calca'
    if any(token in name for token in ['camisa', 'camiseta', 'jaqueta', 'jaleco', 'avental']) or 'corpo' in category:
        return 'tamanho_camisa'
    return None

async def resolve_epi_variation(db, epi: dict, employee: Optional[dict] = None, requested_size: Optional[str] = None, requested_variation_id: Optional[str] = None, require_stock: bool = True) -> Optional[dict]:
    variations = await get_epi_variations(db, epi)
    if requested_variation_id:
        for variation in variations:
            if str(variation['_id']) == requested_variation_id:
                if not require_stock or (variation.get('current_stock', 0) or 0) > 0:
                    return variation
                return None
    desired_size = normalize_size_value(requested_size)
    size_field = infer_employee_size_field(epi)
    if not desired_size and employee and size_field:
        desired_size = normalize_size_value(employee.get(size_field))
    active = [item for item in variations if (item.get('status') or 'ativo') != 'inativo']
    candidates = [item for item in active if not require_stock or (item.get('current_stock', 0) or 0) > 0]
    if desired_size:
        exact = [item for item in candidates if normalize_size_value(item.get('size')) == desired_size]
        if exact:
            return choose_primary_variation(exact)
        if require_stock:
            return None
    generic = [item for item in candidates if not normalize_size_value(item.get('size'))]
    if generic:
        return choose_primary_variation(generic)
    return choose_primary_variation(candidates)

async def get_available_variations_for_delivery(db, epi: dict, require_stock: bool = True) -> List[dict]:
    variations = await get_epi_variations(db, epi)
    active = [item for item in variations if (item.get('status') or 'ativo') != 'inativo']
    return [
        item for item in active
        if not require_stock or (item.get('current_stock', 0) or 0) > 0
    ]

async def get_available_group_variations_for_delivery(db, group_key: str, require_stock: bool = True) -> List[dict]:
    variations = []
    for epi in await get_epis_by_group_key(db, group_key):
        variations.extend(await get_available_variations_for_delivery(db, epi, require_stock=require_stock))
    return variations

def choose_suggested_group_variation(variations: List[dict], requested_size: Optional[str] = None) -> Optional[dict]:
    if not variations:
        return None
    desired_size = normalize_size_value(requested_size)
    if desired_size:
        exact = [item for item in variations if normalize_size_value(item.get('size')) == desired_size]
        if exact:
            return choose_primary_variation(exact)
    generic = [item for item in variations if not normalize_size_value(item.get('size'))]
    return choose_primary_variation(generic or variations)

async def build_delivery_suggestions_for_kit(db, employee: dict, kit: dict) -> dict:
    items = []
    processed_groups = set()
    for raw_item in kit.get('items', []):
        group_key = raw_item.get('epi_group_key') or raw_item.get('epi_base_key')
        epi_base_id = raw_item.get('epi_base_id') or raw_item.get('epi_id')
        epi = None
        if not group_key and epi_base_id and ObjectId.is_valid(str(epi_base_id)):
            epi = await db.epis.find_one({"_id": ObjectId(epi_base_id)})
            if epi:
                group_key = get_epi_group_key(epi)
        if not group_key or group_key in processed_groups:
            continue
        processed_groups.add(group_key)

        summary = await build_epi_group_summary(db, group_key, raw_item)
        if not summary:
            continue
        group_epis = await get_epis_by_group_key(db, group_key)
        representative_epi = epi or (group_epis[0] if group_epis else None)
        size_field = infer_employee_size_field(summary)
        requested_size = employee.get(size_field) if size_field else None
        available_variations = await get_available_group_variations_for_delivery(db, group_key, require_stock=True)
        suggested_variation = choose_suggested_group_variation(available_variations, requested_size)
        selected_variation = suggested_variation or (available_variations[0] if len(available_variations) == 1 else None)
        if available_variations:
            items.append({
                "epi_group_key": group_key,
                "epi_base_key": group_key,
                "epi_base_id": None,
                "epi_id": str(selected_variation.get('epi_id')) if selected_variation else (str(representative_epi['_id']) if representative_epi else None),
                "epi_variation_id": str(selected_variation['_id']) if selected_variation and len(available_variations) == 1 else None,
                "name": summary.get('name'),
                "description": summary.get('description'),
                "category": summary.get('category'),
                "type_category": summary.get('type_category'),
                "quantity": raw_item.get('quantity', 1),
                "size_source": size_field,
                "requested_size": requested_size,
                "status": "available",
                "message": "VariaÃ§Ã£o encontrada em estoque",
                "requires_variation_selection": len(available_variations) > 1,
                "suggested_variation_id": str(suggested_variation['_id']) if suggested_variation else None,
                "available_variations": [serialize_epi_variation(variation) for variation in available_variations],
                "variation": serialize_epi_variation(selected_variation) if selected_variation else None
            })
        else:
            items.append({
                "epi_group_key": group_key,
                "epi_base_key": group_key,
                "epi_base_id": None,
                "epi_id": str(representative_epi['_id']) if representative_epi else None,
                "name": summary.get('name'),
                "description": summary.get('description'),
                "category": summary.get('category'),
                "type_category": summary.get('type_category'),
                "quantity": raw_item.get('quantity', 1),
                "size_source": size_field,
                "requested_size": requested_size,
                "status": "unavailable",
                "message": "Sem estoque disponÃ­vel para tamanho solicitado."
            })
    return {
        "employee_id": str(employee['_id']),
        "employee_name": employee.get('full_name', ''),
        "department": employee.get('department'),
        "kit_id": str(kit['_id']),
        "kit_name": kit.get('name'),
        "items": items
    }

def calculate_epi_status(epi):
    """Calcula status de estoque e validade do EPI"""
    stock_status = 'ok'
    validity_status = 'ok'
    
    # Status de estoque
    if epi.get('current_stock', 0) <= 0:
        stock_status = 'out'
    elif epi.get('current_stock', 0) <= epi.get('min_stock', 0):
        stock_status = 'low'
    
    # Status de validade
    validity_date = ensure_aware_datetime(epi.get('validity_date') or epi.get('ca_validity'))
    if validity_date:
        now = datetime.now(timezone.utc)
        days_until_expiry = (validity_date - now).days
        
        if days_until_expiry < 0:
            validity_status = 'expired'
        elif days_until_expiry <= 30:
            validity_status = 'expiring'
    
    return stock_status, validity_status

@api_router.get('/epis', response_model=List[EPIResponse])
async def get_epis(current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho', 'almoxarifado'))):
    db = await get_db()
    epis = await db.epis.find({}).to_list(1000)
    result = []
    for e in epis:
        result.append(EPIResponse(**(await build_epi_response(db, e))))
    return result

@api_router.post('/epis', response_model=EPIResponse, status_code=status.HTTP_201_CREATED)
async def create_epi(epi_data: EPICreate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    payload = epi_data.model_dump()
    variations_payload = payload.pop('variations', []) or []
    category = payload.get('category') or payload.get('type_category')
    nbr = payload.get('nbr') or payload.get('nbr_number')
    epi_group_key, epi_group_name = await resolve_epi_group_fields(db, payload, payload.get('name'))

    if payload.get('obrigatorio_ca', True):
        if not variations_payload and not payload.get('ca_number'):
            raise HTTPException(status_code=400, detail='Cadastre ao menos uma varia??o com CA para este EPI.')
        if variations_payload and not any(normalize_text(item.get('ca') or item.get('ca_number')) for item in variations_payload):
            raise HTTPException(status_code=400, detail='Cadastre ao menos uma varia??o com CA para este EPI.')

    now = datetime.now(timezone.utc)
    new_epi = {
        'name': payload['name'],
        'category': category,
        'type_category': category,
        'description': payload.get('description'),
        'obrigatorio_ca': payload.get('obrigatorio_ca', True),
        'nbr': nbr,
        'nbr_number': nbr,
        'possui_variacao_tamanho': payload.get('possui_variacao_tamanho', False),
        'epi_group_key': epi_group_key,
        'epi_group_name': epi_group_name,
        'min_stock': payload.get('min_stock', 0),
        'max_stock': payload.get('max_stock'),
        'replacement_period': str(payload['replacement_period'].value) if payload.get('replacement_period') else None,
        'replacement_days': payload.get('replacement_days'),
        'created_by': current_user['id'],
        'created_at': now,
        'updated_at': now
    }
    result = await db.epis.insert_one(new_epi)
    new_epi['_id'] = result.inserted_id

    if not variations_payload and any([
        normalize_text(payload.get('brand')),
        normalize_text(payload.get('model')),
        normalize_text(payload.get('ca_number')),
        payload.get('current_stock', 0) > 0
    ]):
        variations_payload = [payload]

    for raw_variation in variations_payload:
        supplier_id = raw_variation.get('supplier_id')
        quantity = get_variation_stock_quantity(raw_variation)
        variation = {
            'epi_id': str(result.inserted_id),
            'brand': raw_variation.get('marca') or raw_variation.get('brand'),
            'model': raw_variation.get('modelo') or raw_variation.get('model'),
            'ca_number': raw_variation.get('ca') or raw_variation.get('ca_number'),
            'supplier_id': supplier_id,
            'supplier_name': raw_variation.get('fornecedor') or await get_supplier_name(db, supplier_id),
            'ca_validity': raw_variation.get('validade_ca') or raw_variation.get('ca_validity'),
            'size': raw_variation.get('tamanho') or raw_variation.get('size'),
            'color': raw_variation.get('cor') or raw_variation.get('color'),
            'current_stock': quantity,
            'unit_price': raw_variation.get('valor_unitario', raw_variation.get('unit_price')),
            'purchase_date': raw_variation.get('data_compra') or raw_variation.get('purchase_date'),
            'batch': raw_variation.get('lote') or raw_variation.get('batch'),
            'status': raw_variation.get('status') or 'ativo',
            'qr_code': raw_variation.get('qr_code'),
            'internal_code': raw_variation.get('internal_code'),
            'invoice_number': raw_variation.get('invoice_number'),
            'validity_date': raw_variation.get('validity_date'),
            'material': raw_variation.get('material'),
            'technical_standard': raw_variation.get('technical_standard'),
            'quantity_purchased': parse_int_safe(raw_variation.get('quantity_purchased'), quantity),
            'created_at': now,
            'updated_at': now
        }
        await db.epi_variations.insert_one(variation)

    await sync_epi_stock_summary(db, str(result.inserted_id))
    refreshed = await db.epis.find_one({'_id': result.inserted_id})
    return EPIResponse(**(await build_epi_response(db, refreshed)))

@api_router.get('/epis/{epi_id}', response_model=EPIResponse)
async def get_epi(epi_id: str, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho', 'almoxarifado'))):
    db = await get_db()
    epi = await db.epis.find_one({'_id': ObjectId(epi_id)})
    if not epi:
        raise HTTPException(status_code=404, detail='EPI n?o encontrado')
    return EPIResponse(**(await build_epi_response(db, epi)))

@api_router.patch('/epis/{epi_id}', response_model=EPIResponse)
async def update_epi(epi_id: str, epi_data: EPIUpdate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    existing = await db.epis.find_one({'_id': ObjectId(epi_id)})
    if not existing:
        raise HTTPException(status_code=404, detail='EPI n?o encontrado')

    update_data = {k: v for k, v in epi_data.model_dump(exclude_unset=True).items() if k != 'variations'}
    update_data['updated_at'] = datetime.now(timezone.utc)

    if 'category' in update_data or 'type_category' in update_data:
        category = update_data.get('category') or update_data.get('type_category')
        update_data['category'] = category
        update_data['type_category'] = category
    if 'nbr' in update_data or 'nbr_number' in update_data:
        nbr = update_data.get('nbr') or update_data.get('nbr_number')
        update_data['nbr'] = nbr
        update_data['nbr_number'] = nbr
    if 'epi_group_key' in update_data or 'epi_group_name' in update_data:
        group_key, group_name = await resolve_epi_group_fields(
            db,
            {**existing, **update_data},
            update_data.get('name') or existing.get('name')
        )
        update_data['epi_group_key'] = group_key
        update_data['epi_group_name'] = group_name
    if update_data.get('replacement_period'):
        update_data['replacement_period'] = str(update_data['replacement_period'].value) if hasattr(update_data['replacement_period'], 'value') else str(update_data['replacement_period'])

    await db.epis.update_one({'_id': ObjectId(epi_id)}, {'$set': update_data})

    if epi_data.variations is not None:
        await db.epi_variations.delete_many({'epi_id': epi_id})
        now = datetime.now(timezone.utc)
        for raw_variation in epi_data.variations:
            supplier_id = raw_variation.get('supplier_id')
            quantity = get_variation_stock_quantity(raw_variation)
            variation = {
                'epi_id': epi_id,
                'brand': raw_variation.get('marca') or raw_variation.get('brand'),
                'model': raw_variation.get('modelo') or raw_variation.get('model'),
                'ca_number': raw_variation.get('ca') or raw_variation.get('ca_number'),
                'supplier_id': supplier_id,
                'supplier_name': raw_variation.get('fornecedor') or await get_supplier_name(db, supplier_id),
                'ca_validity': raw_variation.get('validade_ca') or raw_variation.get('ca_validity'),
                'size': raw_variation.get('tamanho') or raw_variation.get('size'),
                'color': raw_variation.get('cor') or raw_variation.get('color'),
                'current_stock': quantity,
                'unit_price': raw_variation.get('valor_unitario', raw_variation.get('unit_price')),
                'purchase_date': raw_variation.get('data_compra') or raw_variation.get('purchase_date'),
                'batch': raw_variation.get('lote') or raw_variation.get('batch'),
                'status': raw_variation.get('status') or 'ativo',
                'qr_code': raw_variation.get('qr_code'),
                'internal_code': raw_variation.get('internal_code'),
                'invoice_number': raw_variation.get('invoice_number'),
                'validity_date': raw_variation.get('validity_date'),
                'material': raw_variation.get('material'),
                'technical_standard': raw_variation.get('technical_standard'),
                'quantity_purchased': parse_int_safe(raw_variation.get('quantity_purchased'), quantity),
                'created_at': now,
                'updated_at': now
            }
            await db.epi_variations.insert_one(variation)

    await sync_epi_stock_summary(db, epi_id)
    refreshed = await db.epis.find_one({'_id': ObjectId(epi_id)})
    return EPIResponse(**(await build_epi_response(db, refreshed)))

@api_router.delete('/epis/{epi_id}')
async def delete_epi(epi_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    await db.epi_variations.delete_many({"epi_id": epi_id})
    result = await db.epis.delete_one({"_id": ObjectId(epi_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='EPI não encontrado')
    return {'message': 'EPI excluído'}

# ===================== KITS =====================

async def normalize_kit_items_for_storage(db, kit_items: List[KitItemInput]) -> List[dict]:
    normalized_items = []
    seen_groups = set()
    for item in kit_items:
        group_key = item.epi_group_key
        representative_id = item.epi_base_id or item.epi_id
        if not group_key and representative_id and ObjectId.is_valid(str(representative_id)):
            epi = await db.epis.find_one({"_id": ObjectId(representative_id)})
            if epi:
                group_key = get_epi_group_key(epi)
        if not group_key or group_key in seen_groups:
            continue
        seen_groups.add(group_key)
        normalized_items.append({
            'epi_group_key': group_key,
            'epi_base_key': group_key,
            'name': item.name,
            'description': item.description,
            'category': item.category or item.type_category,
            'type_category': item.type_category or item.category,
            'model': item.model,
            'brand': item.brand,
            'quantity': item.quantity
        })
    return normalized_items

@api_router.get('/kits', response_model=List[KitResponse])
async def get_kits(current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho', 'almoxarifado'))):
    db = await get_db()
    kits = await db.kits.find({}).to_list(1000)
    return [KitResponse(**(await build_kit_response(db, k))) for k in kits]

@api_router.post('/kits', response_model=KitResponse, status_code=status.HTTP_201_CREATED)
async def create_kit(kit_data: KitCreate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    items = await normalize_kit_items_for_storage(db, kit_data.items)
    new_kit = {
        'name': kit_data.name,
        'description': kit_data.description,
        'sector': kit_data.sector,
        'is_mandatory': kit_data.is_mandatory,
        'items': items,
        'created_at': datetime.now(timezone.utc),
        'updated_at': datetime.now(timezone.utc)
    }
    result = await db.kits.insert_one(new_kit)
    new_kit['_id'] = result.inserted_id
    return KitResponse(**(await build_kit_response(db, new_kit)))

@api_router.get('/kits/{kit_id}', response_model=KitResponse)
async def get_kit(kit_id: str, current_user: dict = Depends(get_current_user)):
    db = await get_db()
    kit = await db.kits.find_one({'_id': ObjectId(kit_id)})
    if not kit:
        raise HTTPException(status_code=404, detail='Kit n?o encontrado')
    return KitResponse(**(await build_kit_response(db, kit)))

@api_router.patch('/kits/{kit_id}', response_model=KitResponse)
async def update_kit(kit_id: str, kit_data: KitUpdate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    update_data = {}
    if kit_data.name is not None:
        update_data['name'] = kit_data.name
    if kit_data.description is not None:
        update_data['description'] = kit_data.description
    if kit_data.sector is not None:
        update_data['sector'] = kit_data.sector
    if kit_data.is_mandatory is not None:
        update_data['is_mandatory'] = kit_data.is_mandatory
    if kit_data.items is not None:
        update_data['items'] = await normalize_kit_items_for_storage(db, kit_data.items)
    update_data['updated_at'] = datetime.now(timezone.utc)
    result = await db.kits.find_one_and_update({'_id': ObjectId(kit_id)}, {'$set': update_data}, return_document=True)
    if not result:
        raise HTTPException(status_code=404, detail='Kit n?o encontrado')
    return KitResponse(**(await build_kit_response(db, result)))

@api_router.delete('/kits/{kit_id}')
async def delete_kit(kit_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    result = await db.kits.delete_one({"_id": ObjectId(kit_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Kit não encontrado')
    return {'message': 'Kit excluído'}

# ===================== DELIVERIES =====================

@api_router.post('/deliveries', response_model=DeliveryResponse, status_code=status.HTTP_201_CREATED)
async def create_delivery(delivery_data: DeliveryCreate, current_user: dict = Depends(get_current_user)):
    if not can_deliver_epi(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permiss?o para realizar entregas')

    db = await get_db()
    employee = await db.employees.find_one({'_id': ObjectId(delivery_data.employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador n?o encontrado')
    if not employee.get('photo_path'):
        raise HTTPException(status_code=400, detail='Colaborador n?o possui foto cadastrada. Procure o RH para cadastrar.')

    items_list = []
    for item in delivery_data.items:
        if item.quantity < 1:
            raise HTTPException(status_code=400, detail='Quantidade deve ser maior ou igual a 1.')

        item_dict = item.model_dump()
        effective_epi_id = item.epi_id
        preselected_variation = None
        if item.epi_variation_id and ObjectId.is_valid(str(item.epi_variation_id)):
            preselected_variation = await db.epi_variations.find_one({'_id': ObjectId(item.epi_variation_id)})
            if preselected_variation:
                effective_epi_id = str(preselected_variation.get('epi_id') or effective_epi_id)

        if effective_epi_id:
            epi = await db.epis.find_one({'_id': ObjectId(effective_epi_id)})
            if not epi:
                raise HTTPException(status_code=404, detail='EPI n?o encontrado para entrega.')

            available_variations = await get_available_variations_for_delivery(
                db,
                epi,
                require_stock=not delivery_data.is_return
            )
            if not item.epi_variation_id and len(available_variations) > 1:
                raise HTTPException(status_code=400, detail='Selecione a variacao do EPI antes de concluir a entrega.')

            variation = await resolve_epi_variation(
                db,
                epi,
                employee=employee,
                requested_size=item.size,
                requested_variation_id=item.epi_variation_id,
                require_stock=not delivery_data.is_return
            )
            if not variation:
                raise HTTPException(status_code=400, detail='Sem estoque dispon?vel para tamanho solicitado.')

            if not delivery_data.is_return and (variation.get('current_stock', 0) or 0) < item.quantity:
                raise HTTPException(status_code=400, detail='Estoque insuficiente para a quantidade solicitada.')

            stock_change = item.quantity if delivery_data.is_return else -item.quantity
            await db.epi_variations.update_one({'_id': variation['_id']}, {'$inc': {'current_stock': stock_change}})
            await sync_epi_stock_summary(db, effective_epi_id)

            movement = {
                'movement_type': 'return' if delivery_data.is_return else 'delivery',
                'epi_id': effective_epi_id,
                'epi_group_key': get_epi_group_key(epi),
                'epi_variation_id': str(variation['_id']),
                'quantity': item.quantity if delivery_data.is_return else -item.quantity,
                'employee_id': delivery_data.employee_id,
                'created_by': current_user['id'],
                'created_at': datetime.now(timezone.utc)
            }
            await db.stock_movements.insert_one(movement)

            item_dict.update({
                'epi_name': epi.get('name'),
                'epi_id': effective_epi_id,
                'epi_group_key': get_epi_group_key(epi),
                'epi_variation_id': str(variation['_id']),
                'brand': variation.get('brand'),
                'model': variation.get('model'),
                'size': variation.get('size') or item.size,
                'batch': variation.get('batch') or item.batch,
                'qr_code': variation.get('qr_code') or item.qr_code,
                'ca_number': variation.get('ca_number'),
                'ca_validity': variation.get('ca_validity'),
                'validity_date': variation.get('validity_date'),
                'supplier_id': variation.get('supplier_id'),
                'supplier_name': variation.get('supplier_name'),
                'quantity': item.quantity,
                'kit_id': item.kit_id
            })
            if item.kit_id:
                kit_doc = await db.kits.find_one({'_id': ObjectId(item.kit_id)})
                item_dict['kit_name'] = kit_doc.get('name') if kit_doc else None

        items_list.append(item_dict)

    new_delivery = {
        'employee_id': delivery_data.employee_id,
        'employee_name': employee['full_name'],
        'delivery_type': delivery_data.delivery_type,
        'is_return': delivery_data.is_return,
        'facial_match_score': delivery_data.facial_match_score,
        'facial_photo_path': delivery_data.facial_photo_path,
        'facial_validation_status': delivery_data.facial_validation_status,
        'facial_validation_message': delivery_data.facial_validation_message,
        'facial_liveness_status': delivery_data.facial_liveness_status,
        'facial_second_capture_used': delivery_data.facial_second_capture_used,
        'notes': delivery_data.notes,
        'items': items_list,
        'delivered_by': current_user['id'],
        'delivered_by_name': current_user['username'],
        'created_at': datetime.now(timezone.utc)
    }
    result = await db.deliveries.insert_one(new_delivery)
    new_delivery['_id'] = result.inserted_id
    return DeliveryResponse(**doc_to_response(new_delivery))

@api_router.post('/deliveries/save-photo')
async def save_delivery_photo(
    employee_id: str = Form(...),
    photo_data: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Salva a foto de confirmação da entrega"""
    if not can_deliver_epi(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão')
    
    try:
        # Decodificar base64
        if ',' in photo_data:
            photo_data = photo_data.split(',')[1]
        
        photo_bytes = base64.b64decode(photo_data)
        
        file_name = f'delivery_{employee_id}_{datetime.now(timezone.utc).timestamp()}.jpg'
        file_path = UPLOAD_DIR / 'deliveries' / file_name
        file_path.parent.mkdir(exist_ok=True, parents=True)
        
        with open(file_path, 'wb') as f:
            f.write(photo_bytes)
        
        return {'photo_path': f'/uploads/deliveries/{file_name}'}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'Erro ao salvar foto: {str(e)}')

@api_router.get('/deliveries', response_model=List[DeliveryResponse])
async def get_deliveries(
    current_user: dict = Depends(get_current_user), 
    employee_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    db = await get_db()
    query = {}
    if employee_id:
        query['employee_id'] = employee_id
    
    if start_date:
        query['created_at'] = query.get('created_at', {})
        query['created_at']['$gte'] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
    
    if end_date:
        query['created_at'] = query.get('created_at', {})
        query['created_at']['$lte'] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    
    deliveries = await db.deliveries.find(query).sort("created_at", -1).to_list(1000)
    return [DeliveryResponse(**doc_to_response(d)) for d in deliveries]

# ===================== STOCK =====================

@api_router.get('/stock/alerts')
async def get_stock_alerts(current_user: dict = Depends(get_current_user)):
    db = await get_db()
    
    # EPIs com estoque baixo
    low_stock = await db.epis.find({"$expr": {"$lte": ["$current_stock", "$min_stock"]}}).to_list(100)
    
    # EPIs com validade próxima (30 dias)
    expiry_date = datetime.now(timezone.utc) + timedelta(days=30)
    expiring_soon = await db.epis.find({
        "$or": [
            {"validity_date": {"$ne": None, "$lte": expiry_date}},
            {"ca_validity": {"$ne": None, "$lte": expiry_date}}
        ]
    }).to_list(100)
    
    return {
        'low_stock': [{'id': str(e['_id']), 'name': e['name'], 'current_stock': e['current_stock'], 'min_stock': e['min_stock']} for e in low_stock],
        'expiring_soon': [{'id': str(e['_id']), 'name': e['name'], 'validity_date': e.get('validity_date') or e.get('ca_validity')} for e in expiring_soon]
    }

@api_router.get('/stock/movements')
async def get_stock_movements(current_user: dict = Depends(get_current_user), epi_id: Optional[str] = None):
    db = await get_db()
    query = {}
    if epi_id:
        query['epi_id'] = epi_id
    movements = await db.stock_movements.find(query).sort("created_at", -1).to_list(500)
    return [doc_to_response(m) for m in movements]

# ===================== LICENSE =====================

@api_router.get('/license', response_model=LicenseResponse)
async def get_license(current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    license_doc = await db.panel_license.find_one({})
    if not license_doc:
        raise HTTPException(status_code=404, detail='Licença não encontrada')
    
    now = datetime.now(timezone.utc)
    expires_at = ensure_aware_datetime(license_doc.get('expires_at'))
    if not expires_at:
        raise HTTPException(status_code=500, detail='Licenca com validade invalida')
    
    days_remaining = max(0, (expires_at - now).days)
    
    return LicenseResponse(
        id=str(license_doc['_id']),
        expires_at=expires_at,
        is_blocked=license_doc.get('is_blocked', False),
        days_remaining=days_remaining
    )

@api_router.post('/license/add-days')
async def add_license_days(request: LicenseAddDaysRequest, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    license_doc = await db.panel_license.find_one({})
    if not license_doc:
        raise HTTPException(status_code=404, detail='Licença não encontrada')
    
    new_expires = license_doc['expires_at'] + timedelta(days=request.days)
    await db.panel_license.update_one({"_id": license_doc['_id']}, {"$set": {"expires_at": new_expires}})
    
    history = {
        "license_id": str(license_doc['_id']),
        "user_id": current_user['id'],
        "days_added": request.days,
        "reason": request.reason,
        "created_at": datetime.now(timezone.utc)
    }
    await db.license_history.insert_one(history)
    
    return {'message': f'{request.days} dias adicionados com sucesso'}

# ===================== ALERTAS DE EPI OBRIGATÓRIO E PERIODICIDADE =====================

def get_replacement_days(epi):
    """Retorna o número de dias para troca baseado na periodicidade"""
    period = epi.get('replacement_period')
    if period == 'weekly':
        return 7
    elif period == 'biweekly':
        return 14
    elif period == 'monthly':
        return 30
    elif period == 'custom':
        return epi.get('replacement_days', 30)
    return None

@api_router.get('/alerts/pending-epis')
async def get_pending_epi_alerts(current_user: dict = Depends(get_current_user)):
    """Retorna alertas de EPIs obrigatórios não entregues aos colaboradores"""
    db = await get_db()
    
    alerts = []
    
    # Buscar todos os colaboradores ativos
    employees = await db.employees.find({"status": "active"}).to_list(5000)
    
    # Buscar todos os kits obrigatórios por setor
    kits = await db.kits.find({"is_mandatory": {"$ne": False}}).to_list(100)
    kits_by_sector = {kit.get('sector', '').lower(): kit for kit in kits if kit.get('sector')}
    
    # Buscar todas as entregas
    deliveries = await db.deliveries.find({"is_return": False}).to_list(10000)
    
    # Agrupar entregas por colaborador
    deliveries_by_employee = {}
    for d in deliveries:
        emp_id = d.get('employee_id')
        if emp_id not in deliveries_by_employee:
            deliveries_by_employee[emp_id] = []
        deliveries_by_employee[emp_id].append(d)
    
    for employee in employees:
        emp_id = str(employee['_id'])
        department = (employee.get('department') or '').lower()
        
        # Verificar se há kit obrigatório para o setor
        kit = kits_by_sector.get(department)
        if not kit:
            continue
        
        # Buscar EPIs já entregues ao colaborador
        emp_deliveries = deliveries_by_employee.get(emp_id, [])
        delivered_group_keys = set()
        for d in emp_deliveries:
            for item in d.get('items', []):
                if item.get('epi_group_key'):
                    delivered_group_keys.add(item.get('epi_group_key'))
                elif item.get('epi_id') and ObjectId.is_valid(str(item.get('epi_id'))):
                    epi_doc = await db.epis.find_one({'_id': ObjectId(item.get('epi_id'))})
                    if epi_doc:
                        delivered_group_keys.add(get_epi_group_key(epi_doc))
        delivered_epi_ids = delivered_group_keys
        
        # Verificar EPIs faltantes do kit
        missing_epis = []
        for kit_item in kit.get('items', []):
            group_key = kit_item.get('epi_group_key') or kit_item.get('epi_base_key')
            epi_id = kit_item.get('epi_base_id') or kit_item.get('epi_id')
            epi_doc = None
            if not group_key and epi_id and ObjectId.is_valid(str(epi_id)):
                epi_doc = await db.epis.find_one({'_id': ObjectId(epi_id)})
                if epi_doc:
                    group_key = get_epi_group_key(epi_doc)
            if group_key and group_key not in delivered_group_keys:
                summary = await build_epi_group_summary(db, group_key, kit_item)
                missing_epis.append({
                    'epi_group_key': group_key,
                    'name': (summary or {}).get('name') or kit_item.get('name') or 'EPI',
                    'ca_number': '',
                    'quantity_required': kit_item.get('quantity', 1)
                })
        
        if missing_epis:
            alerts.append({
                'employee_id': emp_id,
                'employee_name': employee.get('full_name', ''),
                'department': employee.get('department', ''),
                'kit_name': kit.get('name', ''),
                'kit_id': str(kit['_id']),
                'missing_epis': missing_epis,
                'alert_type': 'missing_mandatory_epi',
                'message': f"{employee.get('full_name')} ({employee.get('department', '')}) não possui {len(missing_epis)} EPI(s) obrigatório(s) do kit"
            })
    
    return alerts

@api_router.get('/alerts/replacement-due')
async def get_replacement_due_alerts(current_user: dict = Depends(get_current_user)):
    """Retorna alertas de EPIs com troca periódica vencida"""
    db = await get_db()
    
    alerts = []
    now = datetime.now(timezone.utc)
    
    # Buscar EPIs com periodicidade de troca definida
    epis_with_period = await db.epis.find({
        "replacement_period": {"$ne": None}
    }).to_list(500)
    
    epi_periods = {str(e['_id']): e for e in epis_with_period}
    
    if not epi_periods:
        return alerts
    
    # Buscar colaboradores ativos
    employees = await db.employees.find({"status": "active"}).to_list(5000)
    emp_map = {str(e['_id']): e for e in employees}
    
    # Buscar entregas mais recentes de EPIs com periodicidade
    epi_ids = list(epi_periods.keys())
    
    # Usar agregação para encontrar última entrega de cada EPI por colaborador
    pipeline = [
        {"$match": {"is_return": False}},
        {"$unwind": "$items"},
        {"$match": {"items.epi_id": {"$in": epi_ids}}},
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": {"employee_id": "$employee_id", "epi_id": "$items.epi_id"},
            "last_delivery": {"$first": "$created_at"},
            "employee_name": {"$first": "$employee_name"}
        }}
    ]
    
    results = await db.deliveries.aggregate(pipeline).to_list(10000)
    
    for result in results:
        emp_id = result['_id']['employee_id']
        epi_id = result['_id']['epi_id']
        last_delivery = ensure_aware_datetime(result.get('last_delivery'))
        
        if epi_id not in epi_periods:
            continue
        
        epi = epi_periods[epi_id]
        replacement_days = get_replacement_days(epi)
        
        if replacement_days is None:
            continue
        
        if not last_delivery:
            continue
        
        due_date = last_delivery + timedelta(days=replacement_days)
        
        if now > due_date:
            days_overdue = (now - due_date).days
            employee = emp_map.get(emp_id, {})
            
            alerts.append({
                'employee_id': emp_id,
                'employee_name': result.get('employee_name', employee.get('full_name', '')),
                'epi_id': epi_id,
                'epi_name': epi.get('name', ''),
                'ca_number': epi.get('ca_number', ''),
                'nbr_number': epi.get('nbr_number', ''),
                'last_delivery_date': last_delivery.isoformat(),
                'replacement_due_date': due_date.isoformat(),
                'days_overdue': days_overdue,
                'replacement_period': epi.get('replacement_period', ''),
                'alert_type': 'replacement_due',
                'message': f"{result.get('employee_name', 'Colaborador')} não realizou retirada de {epi.get('name')} - {days_overdue} dias de atraso"
            })
    
    return alerts

@api_router.get('/alerts/all')
async def get_all_alerts(current_user: dict = Depends(get_current_user)):
    """Retorna todos os alertas consolidados"""
    pending_epis = await get_pending_epi_alerts(current_user)
    replacement_due = await get_replacement_due_alerts(current_user)
    
    return {
        'pending_epis': pending_epis,
        'replacement_due': replacement_due,
        'total_pending_epis': len(pending_epis),
        'total_replacement_due': len(replacement_due),
        'total_alerts': len(pending_epis) + len(replacement_due)
    }

@api_router.get('/alerts/employee/{employee_id}')
async def get_employee_alerts(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Retorna alertas específicos de um colaborador"""
    db = await get_db()
    
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    alerts = {
        'pending_epis': [],
        'replacement_due': []
    }
    
    department = (employee.get('department') or '').lower()
    
    # Buscar kit obrigatório do setor
    kit = await db.kits.find_one({
        "sector": {"$regex": f"^{department}$", "$options": "i"},
        "is_mandatory": {"$ne": False}
    })
    
    if kit:
        # Buscar EPIs já entregues
        deliveries = await db.deliveries.find({
            "employee_id": employee_id,
            "is_return": False
        }).to_list(1000)
        
        delivered_group_keys = set()
        for d in deliveries:
            for item in d.get('items', []):
                if item.get('epi_group_key'):
                    delivered_group_keys.add(item.get('epi_group_key'))
                elif item.get('epi_id') and ObjectId.is_valid(str(item.get('epi_id'))):
                    epi_doc = await db.epis.find_one({'_id': ObjectId(item.get('epi_id'))})
                    if epi_doc:
                        delivered_group_keys.add(get_epi_group_key(epi_doc))
        delivered_epi_ids = delivered_group_keys
        
        # Verificar EPIs faltantes
        for kit_item in kit.get('items', []):
            epi_id = kit_item.get('epi_base_id') or kit_item.get('epi_id')
            if epi_id and epi_id not in delivered_epi_ids:
                alerts['pending_epis'].append({
                    'epi_id': epi_id,
                    'name': kit_item.get('name', 'EPI'),
                    'ca_number': kit_item.get('ca_number', ''),
                    'quantity_required': kit_item.get('quantity', 1),
                    'message': f"EPI obrigatório não entregue: {kit_item.get('name', 'EPI')}"
                })
    
    # Verificar EPIs com troca vencida
    epi_ids_with_period = []
    epis_with_period = await db.epis.find({"replacement_period": {"$ne": None}}).to_list(500)
    epi_periods = {str(e['_id']): e for e in epis_with_period}
    
    if epi_periods:
        now = datetime.now(timezone.utc)
        
        # Buscar última entrega de cada EPI com periodicidade
        pipeline = [
            {"$match": {"employee_id": employee_id, "is_return": False}},
            {"$unwind": "$items"},
            {"$match": {"items.epi_id": {"$in": list(epi_periods.keys())}}},
            {"$sort": {"created_at": -1}},
            {"$group": {
                "_id": "$items.epi_id",
                "last_delivery": {"$first": "$created_at"}
            }}
        ]
        
        results = await db.deliveries.aggregate(pipeline).to_list(500)
        
        for result in results:
            epi_id = result['_id']
            last_delivery = ensure_aware_datetime(result.get('last_delivery'))
            
            if epi_id not in epi_periods:
                continue
            
            epi = epi_periods[epi_id]
            replacement_days = get_replacement_days(epi)
            
            if replacement_days is None:
                continue
            
            if not last_delivery:
                continue
            
            due_date = last_delivery + timedelta(days=replacement_days)
            
            if now > due_date:
                days_overdue = (now - due_date).days
                alerts['replacement_due'].append({
                    'epi_id': epi_id,
                    'epi_name': epi.get('name', ''),
                    'last_delivery_date': last_delivery.isoformat(),
                    'replacement_due_date': due_date.isoformat(),
                    'days_overdue': days_overdue,
                    'message': f"Troca vencida há {days_overdue} dias: {epi.get('name', '')}"
                })
    
    alerts['total_alerts'] = len(alerts['pending_epis']) + len(alerts['replacement_due'])
    alerts['kit_name'] = kit.get('name') if kit else None
    alerts['kit_id'] = str(kit['_id']) if kit else None
    
    return alerts

# ===================== SETOR-KIT VINCULAÇÃO =====================

@api_router.get('/employees/{employee_id}/delivery-suggestions', response_model=DeliverySuggestionResponse)
async def get_delivery_suggestions(employee_id: str, current_user: dict = Depends(get_current_user), kit_id: Optional[str] = None):
    db = await get_db()
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador nÃ£o encontrado')

    kit = None
    if kit_id:
        kit = await db.kits.find_one({"_id": ObjectId(kit_id)})
    elif employee.get('department'):
        kit = await db.kits.find_one({
            "sector": {"$regex": f"^{employee.get('department')}$", "$options": "i"},
            "is_mandatory": {"$ne": False}
        })

    if not kit:
        return DeliverySuggestionResponse(
            employee_id=employee_id,
            employee_name=employee.get('full_name', ''),
            department=employee.get('department'),
            kit_id=None,
            kit_name=None,
            items=[]
        )

    return DeliverySuggestionResponse(**(await build_delivery_suggestions_for_kit(db, employee, kit)))

@api_router.get('/kits/by-sector/{sector_name}')
async def get_kit_by_sector(sector_name: str, current_user: dict = Depends(get_current_user)):
    """Retorna o kit obrigatório vinculado a um setor"""
    db = await get_db()
    
    kit = await db.kits.find_one({
        "sector": {"$regex": f"^{sector_name}$", "$options": "i"},
        "is_mandatory": {"$ne": False}
    })
    
    if not kit:
        return None

    return KitResponse(**(await build_kit_response(db, kit)))

@api_router.get('/sectors/list')
async def get_sectors_list(current_user: dict = Depends(get_current_user)):
    """Retorna lista de setores com seus kits vinculados"""
    db = await get_db()
    
    # Buscar setores dos colaboradores
    pipeline = [
        {"$match": {"department": {"$ne": None, "$ne": ""}}},
        {"$group": {"_id": "$department"}},
        {"$sort": {"_id": 1}}
    ]
    
    sectors_result = await db.employees.aggregate(pipeline).to_list(100)
    sectors = [s['_id'] for s in sectors_result if s['_id']]
    
    # Buscar kits
    kits = await db.kits.find({}).to_list(100)
    kits_by_sector = {(k.get('sector') or '').lower(): k for k in kits}
    
    result = []
    for sector in sectors:
        kit = kits_by_sector.get(sector.lower())
        result.append({
            'sector_name': sector,
            'has_kit': kit is not None,
            'kit_id': str(kit['_id']) if kit else None,
            'kit_name': kit.get('name') if kit else None,
            'kit_items_count': len(kit.get('items', [])) if kit else 0
        })
    
    return result

# ===================== DASHBOARD =====================

@api_router.get('/dashboard/stats')
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    db = await get_db()
    
    active_employees = await db.employees.count_documents({"status": "active"})
    total_epis = await db.epis.count_documents({})
    low_stock_count = await db.epis.count_documents({"$expr": {"$lte": ["$current_stock", "$min_stock"]}})
    active_employee_docs = await db.employees.find({"status": "active"}).to_list(5000)
    biometric_summaries = await get_biometric_summaries(db, active_employee_docs)
    biometric_registered = sum(1 for item in biometric_summaries.values() if item["status"] == "registered")
    biometric_incomplete = sum(1 for item in biometric_summaries.values() if item["status"] == "incomplete")
    biometric_missing = active_employees - biometric_registered - biometric_incomplete
    biometric_coverage = round((biometric_registered / active_employees) * 100, 1) if active_employees else 0
    
    thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
    recent_deliveries = await db.deliveries.count_documents({
        "is_return": False,
        "created_at": {"$gte": thirty_days_ago}
    })
    
    # EPIs com validade próxima
    expiry_date = datetime.now(timezone.utc) + timedelta(days=30)
    expiring_count = await db.epis.count_documents({
        "$or": [
            {"validity_date": {"$ne": None, "$lte": expiry_date}},
            {"ca_validity": {"$ne": None, "$lte": expiry_date}}
        ]
    })
    
    # NOVO: Contagem de alertas de EPIs obrigatórios pendentes
    pending_epi_alerts = 0
    replacement_due_alerts = 0
    try:
        # Contar alertas de EPIs pendentes (simplificado para performance)
        kits_with_sectors = await db.kits.count_documents({
            "sector": {"$ne": None, "$ne": ""},
            "is_mandatory": {"$ne": False}
        })
        
        # Contar EPIs com periodicidade vencida
        epis_with_period = await db.epis.count_documents({
            "replacement_period": {"$ne": None}
        })
        
        # Se houver kits ou EPIs com periodicidade, buscar alertas completos
        if kits_with_sectors > 0 or epis_with_period > 0:
            all_alerts = await get_all_alerts(current_user)
            pending_epi_alerts = all_alerts.get('total_pending_epis', 0)
            replacement_due_alerts = all_alerts.get('total_replacement_due', 0)
    except Exception as e:
        logger.error(f"Erro ao buscar alertas: {e}")
    
    return {
        'active_employees': active_employees,
        'total_epis': total_epis,
        'low_stock_count': low_stock_count,
        'recent_deliveries': recent_deliveries,
        'expiring_epis': expiring_count,
        'pending_epi_alerts': pending_epi_alerts,
        'replacement_due_alerts': replacement_due_alerts,
        'total_alerts': pending_epi_alerts + replacement_due_alerts,
        'biometric_total_employees': active_employees,
        'biometric_registered': biometric_registered,
        'biometric_missing': biometric_missing,
        'biometric_incomplete': biometric_incomplete,
        'biometric_coverage_percent': biometric_coverage
    }

# ===================== AUTENTICAÇÃO DE FICHA EPI =====================

import hashlib
import secrets

def generate_auth_code(employee_id: str, timestamp: datetime) -> str:
    """Gera código de autenticação único para ficha de EPI"""
    # Criar hash baseado em employee_id, timestamp e secret
    secret_key = os.environ.get('SECRET_KEY', 'cipolatti-secret-key-2026')
    data = f"{employee_id}:{timestamp.isoformat()}:{secret_key}:{secrets.token_hex(4)}"
    hash_obj = hashlib.sha256(data.encode())
    hash_hex = hash_obj.hexdigest()[:10].upper()
    return f"AUT-EPI-{hash_hex}"

@api_router.post('/ficha-auth/generate', response_model=FichaAuthenticationResponse)
async def generate_ficha_authentication(
    data: FichaAuthenticationCreate,
    current_user: dict = Depends(get_current_user)
):
    """Gera código de autenticação para ficha de EPI com validação biométrica"""
    db = await get_db()
    
    # Buscar colaborador
    employee = await db.employees.find_one({"_id": ObjectId(data.employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    # Buscar entregas do colaborador
    delivery_query = {"employee_id": data.employee_id}
    if not data.include_all_history and data.delivery_ids:
        delivery_query["_id"] = {"$in": [ObjectId(d) for d in data.delivery_ids]}
    
    deliveries = await db.deliveries.find(delivery_query).to_list(1000)
    delivery_ids = [str(d["_id"]) for d in deliveries]
    
    # Verificar se há validação biométrica nas entregas
    biometric_validated = False
    max_biometric_score = 0.0
    
    for delivery in deliveries:
        if delivery.get("facial_match_score"):
            biometric_validated = True
            score = delivery.get("facial_match_score", 0)
            if score > max_biometric_score:
                max_biometric_score = score
    
    # Gerar código de autenticação
    now = datetime.now(timezone.utc)
    auth_code = generate_auth_code(data.employee_id, now)
    
    # Dados para QR Code (JSON que pode ser escaneado para verificação)
    qr_data = {
        "type": "EPI_FICHA_AUTH",
        "code": auth_code,
        "employee_id": data.employee_id,
        "employee_name": employee.get("full_name"),
        "date": now.isoformat(),
        "biometric": biometric_validated,
        "verify_url": f"/api/ficha-auth/verify/{auth_code}"
    }
    
    # Salvar no banco
    auth_record = {
        "auth_code": auth_code,
        "employee_id": data.employee_id,
        "employee_name": employee.get("full_name"),
        "delivery_ids": delivery_ids,
        "biometric_validated": biometric_validated,
        "biometric_score": max_biometric_score if biometric_validated else None,
        "created_at": now,
        "created_by": current_user.get("username"),
        "created_by_id": current_user.get("id"),
        "qr_code_data": str(qr_data),
        "status": "active"
    }
    
    result = await db.ficha_authentications.insert_one(auth_record)
    
    return FichaAuthenticationResponse(
        id=str(result.inserted_id),
        auth_code=auth_code,
        employee_id=data.employee_id,
        employee_name=employee.get("full_name"),
        delivery_ids=delivery_ids,
        biometric_validated=biometric_validated,
        biometric_score=max_biometric_score if biometric_validated else None,
        created_at=now,
        created_by=current_user.get("username"),
        qr_code_data=str(qr_data)
    )

@api_router.get('/ficha-auth/verify/{auth_code}', response_model=FichaAuthenticationVerify)
async def verify_ficha_authentication(auth_code: str):
    """Verifica autenticidade de uma ficha de EPI pelo código"""
    db = await get_db()
    
    # Buscar registro de autenticação
    auth_record = await db.ficha_authentications.find_one({"auth_code": auth_code})
    
    if not auth_record:
        return FichaAuthenticationVerify(
            valid=False,
            auth_code=auth_code,
            message="Código de autenticação não encontrado no sistema"
        )
    
    if auth_record.get("status") == "revoked":
        return FichaAuthenticationVerify(
            valid=False,
            auth_code=auth_code,
            employee_name=auth_record.get("employee_name"),
            validation_date=auth_record.get("created_at"),
            message="Este código foi revogado e não é mais válido"
        )
    
    return FichaAuthenticationVerify(
        valid=True,
        auth_code=auth_code,
        employee_name=auth_record.get("employee_name"),
        validation_date=auth_record.get("created_at"),
        biometric_validated=auth_record.get("biometric_validated"),
        message="Documento autêntico - validado pelo sistema GestorEPI"
    )

@api_router.get('/ficha-auth/employee/{employee_id}')
async def get_employee_authentications(
    employee_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas as autenticações de ficha de um colaborador"""
    db = await get_db()
    
    records = await db.ficha_authentications.find(
        {"employee_id": employee_id}
    ).sort("created_at", -1).to_list(100)
    
    return [{
        "id": str(r["_id"]),
        "auth_code": r["auth_code"],
        "created_at": r["created_at"],
        "biometric_validated": r.get("biometric_validated"),
        "status": r.get("status", "active")
    } for r in records]

app.include_router(api_router)
